import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import re
import unicodedata
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from datetime import datetime

def _normalize_text(text):
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r"[^A-Za-z0-9]+", "", text)
    return text.lower().strip()

def create_rich_text(parts):
    """
    Create a CellRichText from a list of (InlineFont, text) pairs.
    Put this at module level so any function can call it.
    """
    rt = CellRichText()
    for font, text in parts:
        rt.append(TextBlock(font, text))
    return rt


def extract_sample_base(identifier):
    if not identifier or not isinstance(identifier, str):
        return ""
    identifier = identifier.strip()
    base = re.sub(r"\s*r\d+(\.\d+)?$", "", identifier, flags=re.IGNORECASE)
    return base.strip()


def extract_run_number(identifier):
    if not identifier or not isinstance(identifier, str):
        return (9999, 0)
    m = re.search(r"r(\d+)(?:\.(\d+))?", identifier, flags=re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (9999, 0)


def _make_fill(hex_color):
    c = hex_color.replace("#", "").upper()
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def _get_valid_co2_rows(rows, col_identifier1):
    valid_indices = []
    seen = {}
    for i, r in enumerate(rows):
        ident = r[col_identifier1 - 1]
        major, minor = extract_run_number(ident)
        if major == 1:
            continue
        if major not in seen:
            seen[major] = (minor, i)
        else:
            prev_minor, prev_i = seen[major]
            if minor < prev_minor:
                seen[major] = (minor, i)
    for _, (_, idx) in seen.items():
        valid_indices.append(idx)
    return sorted(valid_indices)


def add_blue_box(ws):
    """
    Adds the two blue boxes at the lower right (existing) and also formats the
    top big blue area A1:W15 with the requested sub-boxes and values and the
    additional J2:N3 / J4:N8 measured boxes and calculated slope/intercept formulas.
    """

    # Styles
    thick = Side(border_style="thick", color="000000")
    medium = Side(border_style="medium", color="000000")
    blue_fill = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
    black_bold = Font(color="000000", bold=True)
    green_bold = Font(color="008000", bold=True)
    red_bold = Font(color="FF0000", bold=True)
    darkblue_bold = Font(color="000080", bold=True)
    lightblue_bold = Font(color="3399FF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # --- TOP BIG BLUE AREA is assumed filled elsewhere A1:W15 in your script.
    # We'll now put the requested content & boxes inside that area.
    # All coordinates are numeric (row, column) where A=1, B=2, ... W=23

    # A1: put date text "2025-0521"
    today_str = datetime.today().strftime("%Y-%m%d")  # e.g., 2025-1005 for Oct 5, 2025
    ws.cell(row=1, column=1, value=today_str).alignment = Alignment(horizontal="left", vertical="center")

    # C1 heading: bold "Normalization"
    ws.cell(row=1, column=3, value="Normalization").font = black_bold
    ws.cell(row=1, column=3).alignment = center

    # R1 heading: bold "Normalized (vs. VPDB)"
    # Column R is 18
    ws.cell(row=1, column=18, value="Normalized (vs. VPDB)").font = black_bold
    ws.cell(row=1, column=18).alignment = center

    # --- Box around C2:C3 (medium border on outer edges) ---
    c_col = 3
    for r in range(2, 4):  # rows 2-3 inclusive
        cell = ws.cell(row=r, column=c_col)
        cell.fill = blue_fill
        cell.alignment = center
    # apply outer border only
    for r in range(2, 4):
        top = medium if r == 2 else None
        bottom = medium if r == 3 else None
        left = medium
        right = medium
        ws.cell(row=r, column=c_col).border = Border(top=top, bottom=bottom, left=left, right=right)

    # C2: Reference Materials in Bold
    ws.cell(row=2, column=c_col, value="Reference Materials").font = black_bold
    ws.cell(row=2, column=c_col).alignment = center

    # --- Box C4:C8 (outer border) with entries ---
    for r in range(4, 9):
        ws.cell(row=r, column=c_col).fill = blue_fill
        ws.cell(row=r, column=c_col).alignment = center

    # outer border for C4:C8
    for r in range(4, 9):
        top = medium if r == 4 else None
        bottom = medium if r == 8 else None
        left = medium
        right = medium
        ws.cell(row=r, column=c_col).border = Border(top=top, bottom=bottom, left=left, right=right)

    # Fill C5-C8 values (centered, colored) — note: you specified non-bold here
    ws.cell(row=5, column=c_col, value="IAEA 603").font = Font(color="008000", bold=False)
    ws.cell(row=5, column=c_col).alignment = center

    ws.cell(row=6, column=c_col, value="LSVEC").font = Font(color="3399FF", bold=False)
    ws.cell(row=6, column=c_col).alignment = center

    ws.cell(row=7, column=c_col, value="NBS 18").font = Font(color="FF0000", bold=False)
    ws.cell(row=7, column=c_col).alignment = center

    ws.cell(row=8, column=c_col, value="NBS 19").font = Font(color="000080", bold=False)
    ws.cell(row=8, column=c_col).alignment = center

    # --- Box D2:H3 (D=4 .. H=8) outer border; also merge F2:G2 ---
    col_left = 4
    col_right = 8
    row_top = 2
    row_bot = 3

    # fill & center the area
    for r in range(row_top, row_bot + 1):
        for c in range(col_left, col_right + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.alignment = center

    # draw outer border for D2:H3
    for r in range(row_top, row_bot + 1):
        for c in range(col_left, col_right + 1):
            top = thick if r == row_top else None
            bottom = thick if r == row_bot else None
            left = thick if c == col_left else None
            right = thick if c == col_right else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)

    # Merge F2:G2 (F=6, G=7) and write "Published (vs. VPDB)" centered bold
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.cell(row=2, column=6, value="Published (vs. VPDB)").font = black_bold
    ws.cell(row=2, column=6).alignment = center

    # F3 and G3: isotope labels
    ws.cell(row=3, column=6, value="δ¹³C").alignment = center
    ws.cell(row=3, column=7, value="δ¹⁸O").alignment = center

    # --- Box D4:H8 (D=4..H=8 rows 4..8) outer border ---
    rtop = 4
    rbot = 8
    cleft = 4
    cright = 8
    for r in range(rtop, rbot + 1):
        for c in range(cleft, cright + 1):
            ws.cell(row=r, column=c).fill = blue_fill
            ws.cell(row=r, column=c).alignment = center
    # outer border
    for r in range(rtop, rbot + 1):
        for c in range(cleft, cright + 1):
            top = thick if r == rtop else None
            bottom = thick if r == rbot else None
            left = thick if c == cleft else None
            right = thick if c == cright else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)

    # Insert the requested numeric values (with colors & bold)
    # E6: -46.6 in light blue bold (E=5, row=6)
    ws.cell(row=6, column=5, value=-46.6).font = lightblue_bold
    ws.cell(row=6, column=5).alignment = center

    # F5: 2.46 in green bold. (F=6, row=5)
    ws.cell(row=5, column=6, value=2.46).font = green_bold
    ws.cell(row=5, column=6).alignment = center

    # G5: -2.37 green bold (G=7,row=5)
    ws.cell(row=5, column=7, value=-2.37).font = green_bold
    ws.cell(row=5, column=7).alignment = center

    # H6: -26.7 light blue bold (H=8,row=6)
    ws.cell(row=6, column=8, value=-26.7).font = lightblue_bold
    ws.cell(row=6, column=8).alignment = center

    # F7: red -5.01 (F=6,row=7)
    ws.cell(row=7, column=6, value=-5.01).font = red_bold
    ws.cell(row=7, column=6).alignment = center

    # G7: red -23.01
    ws.cell(row=7, column=7, value=-23.01).font = red_bold
    ws.cell(row=7, column=7).alignment = center

    # F8: dark blue bold 1.95
    ws.cell(row=8, column=6, value=1.95).font = darkblue_bold
    ws.cell(row=8, column=6).alignment = center

    # G8: dark blue bold -2.2
    ws.cell(row=8, column=7, value=-2.2).font = darkblue_bold
    ws.cell(row=8, column=7).alignment = center

    # --- NEW REQUESTED CELLS & FORMATTING OUTSIDE/ADJACENT TO THE ABOVE ---

    # InlineFont color hex format: '00RRGGBB' or similar used earlier
    red_if = InlineFont(color='00FF0000', b=True)
    blue_if = InlineFont(color='000000FF', b=True)
    darkblue_if = InlineFont(color='000080', b=True)
    green_if = InlineFont(color='008000', b=True)

    # 1) Cells S2 and T2: δ¹³C
    # Columns: S=19, T=20
    ws.cell(row=2, column=19, value="δ¹³C").alignment = center
    ws.cell(row=2, column=20, value="δ¹³C").alignment = center

    # 2) Cells V2 and W2: δ¹⁸O
    # V=22, W=23
    ws.cell(row=2, column=22, value="δ¹⁸O").alignment = center
    ws.cell(row=2, column=23, value="δ¹⁸O").alignment = center

    # 3) S3 and V3: "18 19" (18 red, 19 blue and bold)
    ws.cell(row=3, column=19).value = create_rich_text([(InlineFont(color='00FF0000', b=False), "18 "), (InlineFont(color='000000FF', b=True), "19")])
    ws.cell(row=3, column=19).alignment = center
    ws.cell(row=3, column=22).value = create_rich_text([(InlineFont(color='00FF0000', b=False), "18 "), (InlineFont(color='000000FF', b=True), "19")])
    ws.cell(row=3, column=22).alignment = center

    # 4) T3 and W3: "18 19 603" (18 red, 19 blue, 603 green) — all bold
    ws.cell(row=3, column=20).value = create_rich_text([(InlineFont(color='00FF0000', b=True), "18 "), (InlineFont(color='000000FF', b=True), "19 "), (InlineFont(color='008000', b=True), "603")])
    ws.cell(row=3, column=20).alignment = center
    ws.cell(row=3, column=23).value = create_rich_text([(InlineFont(color='00FF0000', b=True), "18 "), (InlineFont(color='000000FF', b=True), "19 "), (InlineFont(color='008000', b=True), "603")])
    ws.cell(row=3, column=23).alignment = center

    # 5) R5..R8 labels: R=18
    ws.cell(row=5, column=18, value="IAEA 603").font = Font(color="008000", bold=False)
    ws.cell(row=6, column=18, value="LSVEC").font = Font(color="3399FF", bold=False)
    ws.cell(row=7, column=18, value="NBS 18").font = Font(color="FF0000", bold=False)
    ws.cell(row=8, column=18, value="NBS 19").font = Font(color="000080", bold=False)

    # 6) I10 and I13 content (I=9)
    # I10: "18 19" (18 red bold, 19 dark blue bold)
    ws.cell(row=10, column=9).value = create_rich_text([(InlineFont(color='00FF0000', b=True), "18 "), (InlineFont(color='000080', b=True), "19")])
    ws.cell(row=10, column=9).alignment = center

    # I13: "18 19 603" (18 red bold, 19 blue bold, 603 green bold)
    ws.cell(row=13, column=9).value = create_rich_text([(InlineFont(color='00FF0000', b=True), "18 "), (InlineFont(color='000000FF', b=True), "19 "), (InlineFont(color='008000', b=True), "603")])
    ws.cell(row=13, column=9).alignment = center

    # 7) J10 & J13 should say "Slope" black bold; J11 & J14 "Intercept"
    # J = 10
    ws.cell(row=10, column=10, value="Slope").font = black_bold
    ws.cell(row=11, column=10, value="Intercept").font = black_bold
    ws.cell(row=13, column=10, value="Slope").font = black_bold
    ws.cell(row=14, column=10, value="Intercept").font = black_bold

    # 8) Add another box from J2 to N3, and a box from J4 to N8.
    # J=10, K=11, L=12, M=13, N=14

    # Fill & border for J2:N3
    for r in range(2, 4):
        for c in range(10, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.alignment = center

    for r in range(2, 4):
        for c in range(10, 15):
            top = thick if r == 2 else None
            bottom = thick if r == 3 else None
            left = thick if c == 10 else None
            right = thick if c == 14 else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)

    # Merge J2:N2 and set text "Measured (vs. Working Standard)" centered black bold
    ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=14)
    ws.cell(row=2, column=10, value="Measured (vs. Working Standard)").font = black_bold
    ws.cell(row=2, column=10).alignment = center

    # K3 should have δ¹³C (K=11)
    ws.cell(row=3, column=11, value="δ¹³C").alignment = center

    # N3 should have δ¹⁸O (N=14)
    ws.cell(row=3, column=14, value="δ¹⁸O").alignment = center

    # Fill & border for J4:N8
    for r in range(4, 9):
        for c in range(10, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.alignment = center

    for r in range(4, 9):
        for c in range(10, 15):
            top = thick if r == 4 else None
            bottom = thick if r == 8 else None
            left = thick if c == 10 else None
            right = thick if c == 14 else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)

    # --- Fill K5..K8 (C-avg) and N5..N8 (O-avg) from the precomputed "Average" rows in the sheet ---
    # mapping: reference name (lowercase) -> target row in the J/N box
    ref_to_target_row = {
        "iaea 603": 5,
        "lsvec": 6,
        "nbs 18": 7,
        "nbs 19": 8,
    }

    # initialize found flags to avoid overwriting if multiple average blocks exist
    found_ref = {k: False for k in ref_to_target_row}

    # Scan for the "Average" label that your write_group placed in column R (col 18).
    # The corresponding average values are written one row below that label (avg_row = label_row + 1).
    identifier_col = 3   # column C
    c_avg_col = 18       # column R (C avg stored here)
    o_avg_col = 21       # column U (O avg stored here)

    max_row = ws.max_row
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=c_avg_col).value
        if val and str(val).strip().lower() == "average":
            avg_row = r + 1
            # find the identifier that belongs to this group: look at the last non-empty identifier above the label
            id_row = r - 1
            ident = ""
            # walk upwards until we find a non-empty identifier (limit search to 20 rows to be safe)
            scan_top = max(1, id_row - 20)
            for t in range(id_row, scan_top - 1, -1):
                cellv = ws.cell(row=t, column=identifier_col).value
                if cellv and str(cellv).strip():
                    ident = str(cellv).strip()
                    break

            ident_l = ident.lower()
            ref_key = None
            if "iaea" in ident_l or "603" in ident_l:
                ref_key = "iaea 603"
            elif "lsvec" in ident_l:
                ref_key = "lsvec"
            elif "nbs" in ident_l and "18" in ident_l:
                ref_key = "nbs 18"
            elif "nbs" in ident_l and "19" in ident_l:
                ref_key = "nbs 19"

            # if we detected a reference and haven't filled it yet, write formulas into K/N target cells
            if ref_key and ref_key in ref_to_target_row and not found_ref[ref_key]:
                target_row = ref_to_target_row[ref_key]
                # Write formulas that reference the avg cells (R{avg_row} and U{avg_row}) and round them
                ws.cell(row=target_row, column=11, value=f'=IFERROR(ROUND(R{avg_row},3),"")')  # K = C-avg from R{avg_row}
                ws.cell(row=target_row, column=14, value=f'=IFERROR(ROUND(U{avg_row},3),"")')  # N = O-avg from U{avg_row}

                # apply colors consistent with your sheet styling (not bold)
                if ref_key == "nbs 18":
                    ws.cell(row=target_row, column=11).font = Font(color="FF0000", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="FF0000", bold=False)
                elif ref_key == "nbs 19":
                    ws.cell(row=target_row, column=11).font = Font(color="000080", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="000080", bold=False)
                elif ref_key == "iaea 603":
                    ws.cell(row=target_row, column=11).font = Font(color="008000", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="008000", bold=False)
                elif ref_key == "lsvec":
                    ws.cell(row=target_row, column=11).font = Font(color="3399FF", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="3399FF", bold=False)

                found_ref[ref_key] = True

    # --- Determine numeric rows for columns K and N (5–8) ---
    def get_numeric_rows(ws, col, start=5, end=8):
        rows = []
        for r in range(start, end + 1):
            cell = ws.cell(row=r, column=col)
            val = cell.value
            # Handle numbers, numeric strings, or formulas with cached numeric results
            if isinstance(val, (int, float)):
                rows.append(r)
            elif isinstance(val, str):
                val_str = val.strip()
                # Include formula cells if not empty (i.e., starts with '=' but not just '=IF(...,"")')
                if val_str.startswith('=') and not val_str.upper().endswith('""'):
                    rows.append(r)
                else:
                    try:
                        float(val_str)
                        rows.append(r)
                    except ValueError:
                        continue
        return rows


    k_rows = get_numeric_rows(ws, 11)  # Column K (measured C)
    n_rows = get_numeric_rows(ws, 14)  # Column N (measured O)

    # --- Helper to build Excel range strings ---
    def make_range(letter, rows):
        if not rows:
            return None
        if len(rows) == 1:
            # Duplicate same row so Excel still sees a valid range
            return f"${letter}${rows[0]}:${letter}${rows[0]}"
        return f"${letter}${rows[0]}:${letter}${rows[-1]}"

    # --- Carbon: Published F vs Measured K ---
    if len(k_rows) >= 2:
        f_range = make_range("F", k_rows)
        k_range = make_range("K", k_rows)
        if f_range and k_range:
            ws.cell(row=10, column=11).value = f"=IFERROR(SLOPE({f_range},{k_range}),\"\")"
            ws.cell(row=11, column=11).value = f"=IFERROR(INTERCEPT({f_range},{k_range}),\"\")"
        else:
            ws.cell(row=10, column=11).value = ""
            ws.cell(row=11, column=11).value = ""
    else:
        ws.cell(row=10, column=11).value = ""
        ws.cell(row=11, column=11).value = ""

    # --- Oxygen: Published G vs Measured N ---
    if len(n_rows) >= 2:
        g_range = make_range("G", n_rows)
        n_range = make_range("N", n_rows)
        if g_range and n_range:
            ws.cell(row=10, column=14).value = f"=IFERROR(SLOPE({g_range},{n_range}),\"\")"
            ws.cell(row=11, column=14).value = f"=IFERROR(INTERCEPT({g_range},{n_range}),\"\")"
        else:
            ws.cell(row=10, column=14).value = ""
            ws.cell(row=11, column=14).value = ""
    else:
        ws.cell(row=10, column=14).value = ""
        ws.cell(row=11, column=14).value = ""
    # Format K10,K11,N10,N11 cells to be centered
    ws.cell(row=10, column=11).alignment = center
    ws.cell(row=11, column=11).alignment = center
    ws.cell(row=10, column=14).alignment = center
    ws.cell(row=11, column=14).alignment = center

    # 9) O9, O10, O11 — Aragonite (Kim et al. 2015) and parameters
    # O = 15
    ws.cell(row=9, column=15, value="Aragonite (Kim et al. 2015)").font = green_bold
    ws.cell(row=9, column=15).alignment = center
    ws.cell(row=10, column=15, value=0.992).font = green_bold
    ws.cell(row=10, column=15).alignment = center
    ws.cell(row=11, column=15, value=-16.893).font = green_bold
    ws.cell(row=11, column=15).alignment = center


    try:
        ws.column_dimensions["Z"].width = 11
    except Exception:
        pass


def draw_lower_boxes(ws, divider_top_row, blue_fill, black_bold, green_bold):
    """
    Draws the two lower-right blue boxes positioned just above the divider.
    divider_top_row = first row of the 2-row dark gray divider.
    """
    # Each box will overlap the divider by 2 rows.
    box1_bottom = divider_top_row + 1  # second dark gray line
    box1_top = box1_bottom - 4  # total height 5
    box2_bottom = divider_top_row + 1
    box2_top = box2_bottom - 3  # total height 4

    center = Alignment(horizontal="center", vertical="center")
    thick = Side(style="thick")

    # --- Box 1: Z:AE (cols 26–31)
    for r in range(box1_top, box1_bottom + 1):
        for c in range(26, 32):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.alignment = center
            top = thick if r == box1_top else None
            bottom = thick if r == box1_bottom else None
            left = thick if c == 26 else None
            right = thick if c == 31 else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.cell(box1_top, 26, "Normalized").font = black_bold
    ws.cell(box1_top + 1, 26, "VPDB").font = black_bold
    ws.cell(box1_top + 2, 29, "Calcite").font = black_bold
    ws.cell(box1_top + 2, 30, "Calcite").font = black_bold
    ws.cell(box1_top + 2, 31, "Aragonite").font = green_bold

    ws.cell(box1_top + 3, 26, "δ¹³C")
    ws.cell(box1_top + 3, 27, "δ¹³C")
    ws.cell(box1_top + 3, 29, "δ¹⁸O")
    ws.cell(box1_top + 3, 30, "δ¹⁸O")
    ws.cell(box1_top + 3, 31, "δ¹⁸O")

    red_font = InlineFont(color='00FF0000', b=True)
    blue_font = InlineFont(color='000000FF', b=True)
    green_font = InlineFont(color='008000', b=True)

    ws.cell(box1_top + 4, 26).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    ws.cell(box1_top + 4, 27).value = create_rich_text([(red_font, "18 "), (blue_font, "19 "), (green_font, "603")])
    ws.cell(box1_top + 4, 29).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    ws.cell(box1_top + 4, 30).value = create_rich_text([(red_font, "18 "), (blue_font, "19 "), (green_font, "603")])
    ws.cell(box1_top + 4, 31).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])

    # --- Box 2: AG:AH (cols 33–34)
    for r in range(box2_top, box2_bottom + 1):
        for c in range(33, 35):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.alignment = center
            top = thick if r == box2_top else None
            bottom = thick if r == box2_bottom else None
            left = thick if c == 33 else None
            right = thick if c == 34 else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.cell(box2_top, 33, "VSMOW").font = black_bold
    ws.cell(box2_top + 1, 33, "Calcite").font = black_bold
    ws.cell(box2_top + 1, 34, "Aragonite").font = green_bold
    ws.cell(box2_top + 2, 33, "δ¹⁸O")
    ws.cell(box2_top + 2, 34, "δ¹⁸O")

    ws.cell(box2_top + 3, 33).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    ws.cell(box2_top + 3, 34).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])

    try:
        ws.column_dimensions["Z"].width = 11
    except Exception:
        pass



def step4_group(file_path):
    reference_names = ["CO2", "NBS 18", "NBS 19", "IAEA 603", "LSVEC"]
    ref_set = {_normalize_text(r) for r in reference_names}

    wb = openpyxl.load_workbook(file_path)

    if "Last 6" not in wb.sheetnames:
        raise ValueError("Sheet 'Last 6' not found!")

    ws_last6 = wb["Last 6"]

    # Ensure Group sheet is recreated to the LEFT of "Last 6"
    if "Group" in wb.sheetnames:
        wb.remove(wb["Group"])
    last6_index = wb.sheetnames.index("Last 6")
    ws_group = wb.create_sheet("Group", last6_index)

    # make sure sheets are not grouped/selected together
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    # mark Last 6 as the selected tab (prevents grouping with newly created sheet)
    try:
        ws_last6.sheet_view.tabSelected = True
        ws_group.sheet_view.tabSelected = False
    except Exception:
        pass

    blue_fill = _make_fill("DAE9F8")
    dark_fill = _make_fill("808080")
    gray_fill = _make_fill("E7E7E7")

    color_fonts = {
        "nbs18": Font(color="FF0000"),
        "nbs19": Font(color="000080"),
        "iaea603": Font(color="008000"),
        "lsvec": Font(color="3399FF"),
    }

    for row in ws_group.iter_rows(min_row=1, max_row=15, min_col=1, max_col=23):
        for cell in row:
            cell.fill = blue_fill

    headers = []
    first_row = list(ws_last6[1]) if ws_last6.max_row >= 1 else []
    for col_idx in range(24):
        headers.append(first_row[col_idx].value if col_idx < len(first_row) else None)
        ws_group.cell(row=18, column=col_idx + 1, value=headers[-1])

    data_rows = []
    for row in ws_last6.iter_rows(min_row=2, max_col=24, values_only=True):
        if any(row):
            row = list(row) + [None] * (24 - len(row))
            data_rows.append(tuple(row[:24]))

    col_identifier1 = 3
    groups = {}
    for r in data_rows:
        ident = r[col_identifier1 - 1]
        base = extract_sample_base(ident)
        norm = _normalize_text(base)
        if norm not in groups:
            groups[norm] = {"base": base, "rows": []}
        groups[norm]["rows"].append(r)

    for g in groups.values():
        g["rows"].sort(key=lambda r: extract_run_number(r[col_identifier1 - 1]))

    ref_groups = []
    other_groups = []
    for norm, g in groups.items():
        if norm in ref_set:
            ref_groups.append((norm, g))
        else:
            other_groups.append((norm, g))

    current_row = 19

    # regex to detect "N Arag" or "N. Arag" (optional dot, optional spaces)
    n_arag_re = re.compile(r"\bn\.?\s*arag\b", flags=re.IGNORECASE)

    def write_group(norm, g, is_reference=True):
        nonlocal current_row
        base_name = _normalize_text(g["base"])
        rows = g["rows"]
        start_row = current_row

        font_color = color_fonts.get(base_name)
        valid_indices = []
        if base_name == "co2":
            valid_indices = _get_valid_co2_rows(rows, col_identifier1)

        row_map = []
        for i, row in enumerate(rows):
            excel_row = current_row
            identifier_value = str(row[col_identifier1 - 1] or "")
            for col_idx, val in enumerate(row, start=1):
                cell = ws_group.cell(row=excel_row, column=col_idx, value=val)
                if base_name == "co2" and i in valid_indices:
                    for col in (18, 21):
                        ws_group.cell(row=excel_row, column=col).fill = gray_fill
                    cell.fill = gray_fill
                if font_color and col_idx == col_identifier1:
                    cell.font = font_color
            if base_name == "co2" and i in valid_indices:
                row_map.append(excel_row)
            current_row += 1

        end_row = current_row - 1

        if is_reference:
            # Reference group summary formulas
            for col_offset, label in enumerate(["Average", "Stdev", "Count"], start=0):
                cell = ws_group.cell(row=current_row, column=18 + col_offset, value=label)
                cell.alignment = Alignment(horizontal="right")
                cell2 = ws_group.cell(row=current_row, column=21 + col_offset, value=label)
                cell2.alignment = Alignment(horizontal="right")

            avg_row = current_row + 1
            if base_name == "co2" and row_map:
                r_ranges = ",".join([f"R{r}" for r in row_map])
                u_ranges = ",".join([f"U{r}" for r in row_map])
                ws_group.cell(row=avg_row, column=18, value=f"=ROUND(AVERAGE({r_ranges}),3)")
                ws_group.cell(row=avg_row, column=19, value=f"=ROUND(STDEV({r_ranges}),3)")
                ws_group.cell(row=avg_row, column=20, value=f"=ROUND(COUNT({r_ranges}),3)")
                ws_group.cell(row=avg_row, column=21, value=f"=ROUND(AVERAGE({u_ranges}),3)")
                ws_group.cell(row=avg_row, column=22, value=f"=ROUND(STDEV({u_ranges}),3)")
                ws_group.cell(row=avg_row, column=23, value=f"=ROUND(COUNT({u_ranges}),3)")
            else:
                ws_group.cell(row=avg_row, column=18, value=f"=ROUND(AVERAGE(R{start_row}:R{end_row}),3)")
                ws_group.cell(row=avg_row, column=19, value=f"=ROUND(STDEV(R{start_row}:R{end_row}),3)")
                ws_group.cell(row=avg_row, column=20, value=f"=ROUND(COUNT(R{start_row}:R{end_row}),3)")
                ws_group.cell(row=avg_row, column=21, value=f"=ROUND(AVERAGE(U{start_row}:U{end_row}),3)")
                ws_group.cell(row=avg_row, column=22, value=f"=ROUND(STDEV(U{start_row}:U{end_row}),3)")
                ws_group.cell(row=avg_row, column=23, value=f"=ROUND(COUNT(U{start_row}:U{end_row}),3)")

            if font_color:
                for col in range(18, 24):
                    ws_group.cell(row=current_row, column=col).font = font_color
                    ws_group.cell(row=avg_row, column=col).font = font_color

            current_row += 3

        else:
            # Non-reference groups
            for r in range(start_row, current_row):
                ident_val = str(ws_group.cell(row=r, column=col_identifier1).value or "")

                # If N arag / N. arag → do Z, AC, AE, AH; skip AG; row text green
                if n_arag_re.search(ident_val):
                    # Z (col 26) and AC (col 29) — rounded to 2 dp
                    cell_z = ws_group.cell(row=r, column=26, value=f'=IFERROR(ROUND(($K$10*R{r})+$K$11,2),"")')
                    cell_z.font = Font(bold=True)

                    cell_ac = ws_group.cell(row=r, column=29, value=f'=IFERROR(ROUND(($N$10*U{r})+$N$11,2),"")')
                    cell_ac.font = Font(bold=True)

                    # AE (col 31) and AH (col 34) — rounded to 2 dp
                    cell_ae = ws_group.cell(row=r, column=31, value=f'=IFERROR(ROUND(($O$10*U{r})+$O$11,2),"")')
                    cell_ae.font = Font(bold=True)

                    cell_ah = ws_group.cell(row=r, column=34, value=f'=IFERROR(ROUND((1.03092*AE{r})+30.92,2),"")')
                    cell_ah.font = Font(bold=True)

                    # Clear AG (col 33) to avoid leftovers
                    ws_group.cell(row=r, column=33, value=None)

                    # Make entire row text green (preserve bolding on Z/AC/AE/AH by re-applying)
                    green_font = Font(color="008000")
                    green_bold = Font(color="008000", bold=True)
                    for c in range(1, 36):
                        if c in (26, 29, 31, 34):
                            ws_group.cell(row=r, column=c).font = green_bold
                        else:
                            ws_group.cell(row=r, column=c).font = green_font

                else:
                    # Normal ones: Z, AC, and AG — all rounded to 2 dp
                    cell_z = ws_group.cell(row=r, column=26, value=f'=IFERROR(ROUND(($K$10*R{r})+$K$11,2),"")')
                    cell_z.font = Font(bold=True)

                    cell_ac = ws_group.cell(row=r, column=29, value=f'=IFERROR(ROUND(($N$10*U{r})+$N$11,2),"")')
                    cell_ac.font = Font(bold=True)

                    cell_ag = ws_group.cell(row=r, column=33, value=f'=IFERROR(ROUND((1.03092*AC{r})+30.92,2),"")')
                    cell_ag.font = Font(bold=True)

            current_row += 3

    # Write reference groups first
    for norm, g in ref_groups:
        write_group(norm, g, is_reference=True)

    # Divider
    if ref_groups:
        current_row += 8
        divider_top_row = current_row  # store divider start row

        for _ in range(2):
            for col in range(1, 702):
                ws_group.cell(row=current_row, column=col).fill = dark_fill
            current_row += 1

        for col_idx, h in enumerate(headers, start=1):
            ws_group.cell(row=current_row, column=col_idx, value=h)
        current_row += 1
        draw_lower_boxes(ws_group, divider_top_row, blue_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"))


    # Write non-reference groups
    for norm, g in other_groups:
        write_group(norm, g, is_reference=False)

    # Fill grey cells
    max_row = ws_group.max_row + 50
    for row in range(16, max_row + 1):
        for col in (18, 21):
            ws_group.cell(row=row, column=col).fill = gray_fill

    ws_group.column_dimensions["C"].width = 22
    ws_group.column_dimensions["R"].width = 22

    # --- Call it after filling the groups ---
    add_blue_box(ws_group)

    wb.save(file_path)
    print(f"✅ Step 4: GROUP completed on {file_path}")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.views import Selection

def step1_data(file_path, sheet_name='Default_Gas_Bench.wke'):
    """
    Step 1: DATA
    Reads the Excel file, transforms it (padded rows, formulas, rounding),
    and saves the file.
    """
    new_sheet_name = 'Data'

    # Read original data into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

    # Build headers for the new sheet
    headers = [
        'Line', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis',
        'Preparation', 'Peak Nr', 'Rt', 'Ampl 44', 'Area All',
        'd 13C/12C', 'd 18O/16O',
        '', '', '', '',  # spacer columns
        'C avg', 'C stdev', '', 'O avg', 'O stdev', '',
        'Sum area all', 'area peaks', 'funny peaks', 'min intensity'
    ]

    # Load workbook and remove old sheet if exists
    wb = load_workbook(file_path)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # Create new sheet before the original sheet
    first_index = wb.index(wb[sheet_name])
    ws = wb.create_sheet(new_sheet_name, first_index)

    # Ensure only the new sheet is selected (prevents Excel grouping sheets)
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws.sheet_view.tabSelected = True
    wb.active = wb.index(ws)
    # set a default selection using the Selection object (fixes the TypeError)
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Write header row
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # Add one blank row after headers
    cur_row = 3

    # Build maps:
    # col_map: header (non-empty) -> excel column index
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}

    # Helper: normalize function (collapse whitespace, lower)
    def normalize_name(s):
        if s is None:
            return ''
        return ' '.join(str(s).split()).lower()

    # Map dataframe columns (normalized) -> original df column name
    df_cols_norm = {normalize_name(c): c for c in df.columns}

    # Map header -> matching df column name (if any)
    header_to_dfcol = {}
    for h in headers:
        if not h:
            header_to_dfcol[h] = None
            continue
        nh = normalize_name(h)
        if nh in df_cols_norm:
            header_to_dfcol[h] = df_cols_norm[nh]
            continue
        # try a join-without-space match
        nh_join = nh.replace(' ', '')
        match = None
        for dc_norm, dc in df_cols_norm.items():
            if dc_norm.replace(' ', '') == nh_join:
                match = dc
                break
        if match:
            header_to_dfcol[h] = match
            continue
        # fallback: contains 'ampl' or common tokens
        match = None
        for dc_norm, dc in df_cols_norm.items():
            if nh in dc_norm or dc_norm in nh or nh_join in dc_norm.replace(' ', '') or dc_norm.replace(' ', '').find(nh_join) >= 0:
                match = dc
                break
        header_to_dfcol[h] = match  # may be None if nothing matched

    # locate important excel column indexes (these use the new-sheet headers)
    col_area = col_map.get('Area All')
    col_c = col_map.get('d 13C/12C')
    col_o = col_map.get('d 18O/16O')
    col_ampl = col_map.get('Ampl 44')
    col_funny = col_map.get('funny peaks')
    col_minint = col_map.get('min intensity')

    # Excel column letters (for formula creation) — only for defined columns
    col_letter_area = get_column_letter(col_area) if col_area else None
    col_letter_c = get_column_letter(col_c) if col_c else None
    col_letter_o = get_column_letter(col_o) if col_o else None
    col_letter_ampl = get_column_letter(col_ampl) if col_ampl else None

    # Row labels + spacing rules
    summary_layout = [
        ("ref avg", 0),
        ("all", 3),
        ("last 6", 0),
        ("start", 2),
        ("end", 0),
        ("delta", 0),
    ]

    # Summary column starting positions (with blanks accounted for)
    col_label = 17  # Q
    col_c_avg = col_label + 1
    col_c_stdev = col_label + 2
    col_o_avg = col_label + 4  # skip blank after C stdev
    col_o_stdev = col_label + 5
    col_sum_area = col_label + 7  # skip blank after O stdev

    # Colors
    fill_label = PatternFill(start_color="cdffcc", end_color="cdffcc", fill_type="solid")  # green
    fill_funny_min = PatternFill(start_color="cdfeff", end_color="cdfeff", fill_type="solid")  # blue

    # Track delta rows (to leave the blank row below each delta uncolored)
    all_delta_rows = []

    # Group by Line (preserve order)
    grouped = df.groupby('Line', sort=False)

    for line, group in grouped:
        # insert a spacer row between groups (except before first)
        if cur_row != 3:
            cur_row += 1

        first_data_row = cur_row

        # Build padded_rows: 11 rows — use actual rows when present, otherwise create synthetic rows
        padded_rows = []
        for i in range(11):
            if i < len(group):
                padded_rows.append(group.iloc[i].to_dict())  # keys are df column names
            else:
                # synthetic row: copy A/B/C from first real row (if exists) and set Peak Nr; other df-column keys None
                base_vals = group.iloc[0].to_dict() if len(group) > 0 else {}
                blank_row = {col: None for col in df.columns}
                # copy Line, Time Code, Identifier 1 if present in original df columns
                for name in ['Line', 'Time Code', 'Identifier 1']:
                    if name in blank_row:
                        blank_row[name] = base_vals.get(name)
                # set Peak Nr if 'Peak Nr' exists as a df column
                if 'Peak Nr' in blank_row:
                    blank_row['Peak Nr'] = i + 1
                padded_rows.append(blank_row)

        # Write each padded row into the new sheet
        for row_dict in padded_rows:
            # iterate headers for consistent column placement in new sheet
            for h in headers:
                if not h:
                    continue

                # Skip writing into "Sum area all" during data row writing
                if h == "Sum area all":
                    continue

                excel_col = col_map[h]
                source_col = header_to_dfcol.get(h)
                val = None
                if source_col and source_col in row_dict:
                    val = row_dict.get(source_col)

                cell = ws.cell(row=cur_row, column=excel_col, value=val)

                if h in ["Identifier 2", "Analysis"] and val is not None:
                    cell.number_format = '@'

            cur_row += 1

        last_data_row = cur_row - 1

        # place summary formulas (C & O stats etc.)
        last7_start = max(first_data_row, last_data_row - 6)
        last6_start = max(first_data_row, last_data_row - 5)
        start_of_last6 = first_data_row + 5 if (first_data_row + 5) <= last_data_row else first_data_row

        summary_row = first_data_row
        row_positions = {}

        for label, spacing in summary_layout:
            summary_row += spacing
            row_positions[label] = summary_row

            ws.cell(row=summary_row, column=col_label, value=label)

            # only create formulas if relevant columns exist
            if label == "ref avg" and col_letter_c and col_letter_o:
                idx1, idx2, idx4 = first_data_row, first_data_row + 1, first_data_row + 3
                ws.cell(row=summary_row, column=col_c_avg,
                        value=f"=ROUND(AVERAGE({col_letter_c}{idx1},{col_letter_c}{idx2},{col_letter_c}{idx4}),3)")
                ws.cell(row=summary_row, column=col_c_stdev,
                        value=f"=ROUND(STDEV({col_letter_c}{idx1},{col_letter_c}{idx2},{col_letter_c}{idx4}),3)")
                ws.cell(row=summary_row, column=col_o_avg,
                        value=f"=ROUND(AVERAGE({col_letter_o}{idx1},{col_letter_o}{idx2},{col_letter_o}{idx4}),3)")
                ws.cell(row=summary_row, column=col_o_stdev,
                        value=f"=ROUND(STDEV({col_letter_o}{idx1},{col_letter_o}{idx2},{col_letter_o}{idx4}),3)")

            elif label == "all" and col_letter_c and col_letter_o:
                ws.cell(row=summary_row, column=col_c_avg,
                        value=f"=ROUND(AVERAGE({col_letter_c}{last7_start}:{col_letter_c}{last_data_row}),3)")
                ws.cell(row=summary_row, column=col_c_stdev,
                        value=f"=ROUND(STDEV({col_letter_c}{last7_start}:{col_letter_c}{last_data_row}),3)")
                ws.cell(row=summary_row, column=col_o_avg,
                        value=f"=ROUND(AVERAGE({col_letter_o}{last7_start}:{col_letter_o}{last_data_row}),3)")
                ws.cell(row=summary_row, column=col_o_stdev,
                        value=f"=ROUND(STDEV({col_letter_o}{last7_start}:{col_letter_o}{last_data_row}),3)")
                if col_letter_area:
                    ws.cell(row=summary_row, column=col_sum_area,
                            value=f"=ROUND(SUM({col_letter_area}{last7_start}:{col_letter_area}{last_data_row}),2)")

            elif label == "last 6" and col_letter_c and col_letter_o:
                ws.cell(row=summary_row, column=col_c_avg,
                        value=f"=ROUND(AVERAGE({col_letter_c}{last6_start}:{col_letter_c}{last_data_row}),3)")
                ws.cell(row=summary_row, column=col_c_stdev,
                        value=f"=ROUND(STDEV({col_letter_c}{last6_start}:{col_letter_c}{last_data_row}),3)")
                ws.cell(row=summary_row, column=col_o_avg,
                        value=f"=ROUND(AVERAGE({col_letter_o}{last6_start}:{col_letter_o}{last_data_row}),3)")
                ws.cell(row=summary_row, column=col_o_stdev,
                        value=f"=ROUND(STDEV({col_letter_o}{last6_start}:{col_letter_o}{last_data_row}),3)")
                if col_letter_area:
                    ws.cell(row=summary_row, column=col_sum_area,
                            value=f"=ROUND(SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row}),2)")

            elif label == "start" and col_letter_c and col_letter_o:
                ws.cell(row=summary_row, column=col_c_avg, value=f"=ROUND({col_letter_c}{start_of_last6},3)")
                ws.cell(row=summary_row, column=col_o_avg, value=f"=ROUND({col_letter_o}{start_of_last6},3)")

            elif label == "end" and col_letter_c and col_letter_o:
                ws.cell(row=summary_row, column=col_c_avg, value=f"=ROUND({col_letter_c}{last_data_row},3)")
                second_last_row = last_data_row - 1 if last_data_row > first_data_row else last_data_row
                ws.cell(row=summary_row, column=col_o_avg, value=f"=ROUND({col_letter_o}{second_last_row},3)")

            elif label == "delta" and col_letter_c and col_letter_o:
                start_row = row_positions["start"]
                end_row = row_positions["end"]
                ws.cell(row=summary_row, column=col_c_avg,
                        value=f"=ROUND({get_column_letter(col_c_avg)}{end_row}-{get_column_letter(col_c_avg)}{start_row},3)")
                ws.cell(row=summary_row, column=col_o_avg,
                        value=f"=ROUND({get_column_letter(col_o_avg)}{end_row}-{get_column_letter(col_o_avg)}{start_row},3)")

            summary_row += 1


        all_delta_rows.append(row_positions["delta"])

        # --- Funny peaks & min intensity formulas for this 11-row block ---
        # Only proceed if Ampl column exists and target columns exist
        if col_letter_ampl and col_funny and col_minint:
            for i in range(11):
                row_num = first_data_row + i
                if i < 4:
                    ws.cell(row=row_num, column=col_funny, value="ref")
                    ws.cell(row=row_num, column=col_minint, value=None)
                else:
                    ws.cell(row=row_num, column=col_funny,
                            value=f'=IF({col_letter_ampl}{row_num}>{col_letter_ampl}{row_num+1},IF({col_letter_ampl}{row_num+1}<{col_letter_ampl}{row_num},"ok","check"),"check")')
                    ws.cell(row=row_num, column=col_minint,
                            value=f'=IF({col_letter_ampl}{row_num}<400,"check","ok")')

    # --- Apply coloring patterns ---

    # Column Q (green) — color every cell in column Q except the row immediately after each delta
    skip_rows = {d + 1 for d in all_delta_rows}
    for r in range(1, ws.max_row + 1):
        if r in skip_rows:
            continue
        ws.cell(row=r, column=col_label).fill = fill_label

    # Columns Z & AA (blue) — follow same pattern as column Q
    if col_funny and col_minint:
        for r in range(1, ws.max_row + 1):
            if r in skip_rows:
                continue
            ws.cell(row=r, column=col_funny).fill = fill_funny_min
            ws.cell(row=r, column=col_minint).fill = fill_funny_min

    # Example at the end:
    wb.save(file_path)
    print(f"Step 1: DATA completed on {file_path}")

import os
from copy import copy, deepcopy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell.rich_text import CellRichText, TextBlock

def _is_formula_cell(cell):
    """Return True if the cell is a formula."""
    try:
        if getattr(cell, "data_type", None) == "f":
            return True
        val = cell.value
        return isinstance(val, str) and val.startswith("=")
    except Exception:
        return False

def _try_refresh_with_xlwings(path):
    try:
        import xlwings as xw
    except Exception:
        return False

    app = None
    book = None
    try:
        app = xw.App(visible=False, add_book=False)
        book = app.books.open(os.path.abspath(path))
        try:
            book.app.api.Application.CalculateFull()
        except Exception:
            try:
                book.app.api.Application.Calculate()
            except Exception:
                try:
                    book.app.calculate()
                except Exception:
                    pass
        book.save()
        book.close()
        app.quit()
        return True
    except Exception:
        try:
            if book is not None:
                book.close()
        except Exception:
            pass
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass
        return False

def step5_summary(file_path):
    source_sheet = "Group"
    new_sheet_name = "Summary"

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb_fmt = load_workbook(file_path, data_only=False)
    wb_val = load_workbook(file_path, data_only=True)

    if source_sheet not in wb_fmt.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found.")

    ws_fmt = wb_fmt[source_sheet]
    ws_val = wb_val[source_sheet]

    def _cell_rgb_upper(cell):
        try:
            fg = getattr(cell.fill, "fgColor", None)
            if fg is not None:
                rgb = getattr(fg, "rgb", None)
                if rgb:
                    return str(rgb).upper()
            sc = getattr(cell.fill, "start_color", None)
            if sc is not None:
                rgb2 = getattr(sc, "rgb", None)
                if rgb2:
                    return str(rgb2).upper()
        except Exception:
            pass
        return None

    def _is_gray808080(cell):
        rgb = _cell_rgb_upper(cell)
        return bool(rgb and rgb.endswith("808080"))

    gray_band_start = None
    check_start_col = 26
    check_end_col = 34
    range_width = check_end_col - check_start_col + 1
    threshold = max(1, range_width // 2)

    for r in range(1, ws_fmt.max_row):
        count_r = sum(1 for c in range(check_start_col, check_end_col + 1)
                      if _is_gray808080(ws_fmt.cell(row=r, column=c)))
        count_r1 = sum(1 for c in range(check_start_col, check_end_col + 1)
                       if _is_gray808080(ws_fmt.cell(row=r + 1, column=c)))
        if count_r >= threshold and count_r1 >= threshold:
            gray_band_start = r
            break

    if gray_band_start is None:
        start_col_l = 12
        end_col = ws_fmt.max_column
        width2 = max(1, end_col - start_col_l + 1)
        threshold2 = max(1, width2 // 2)
        for r in range(1, ws_fmt.max_row):
            count_r = sum(1 for c in range(start_col_l, end_col + 1)
                          if _is_gray808080(ws_fmt.cell(row=r, column=c)))
            count_r1 = sum(1 for c in range(start_col_l, end_col + 1)
                           if _is_gray808080(ws_fmt.cell(row=r + 1, column=c)))
            if count_r >= threshold2 and count_r1 >= threshold2:
                gray_band_start = r
                break

    if gray_band_start is None:
        for r in range(1, ws_fmt.max_row):
            any_r = any(_is_gray808080(ws_fmt.cell(row=r, column=c)) for c in range(1, ws_fmt.max_column + 1))
            any_r1 = any(_is_gray808080(ws_fmt.cell(row=r + 1, column=c)) for c in range(1, ws_fmt.max_column + 1)) if r < ws_fmt.max_row else False
            if any_r and any_r1:
                gray_band_start = r
                break

    if gray_band_start is None:
        raise ValueError("Could not find the 2-row dark gray band (color #808080) in 'Group' sheet.")

    start_row = max(1, gray_band_start - 3)
    source_cols = list(range(1, 4)) + list(range(26, 35))

    needs_refresh = False
    for r in range(start_row, min(start_row + 30, ws_fmt.max_row + 1)):
        for c in source_cols:
            src = ws_fmt.cell(row=r, column=c)
            valcell = ws_val.cell(row=r, column=c)
            if _is_formula_cell(src) and (valcell.value is None):
                needs_refresh = True
                break
        if needs_refresh:
            break

    if needs_refresh:
        refreshed = _try_refresh_with_xlwings(file_path)
        if refreshed:
            wb_fmt = load_workbook(file_path, data_only=False)
            wb_val = load_workbook(file_path, data_only=True)
            ws_fmt = wb_fmt[source_sheet]
            ws_val = wb_val[source_sheet]

    if new_sheet_name in wb_fmt.sheetnames:
        del wb_fmt[new_sheet_name]

    ws_new = wb_fmt.create_sheet(new_sheet_name, index=wb_fmt.index(ws_fmt))
    mapping = {src_col: idx for idx, src_col in enumerate(source_cols, start=1)}

    new_row = 1
    for r in range(start_row, ws_fmt.max_row + 1):
        for src_col in source_cols:
            new_col = mapping[src_col]
            src_cell_fmt = ws_fmt.cell(row=r, column=src_col)
            src_cell_val = ws_val.cell(row=r, column=src_col)
            dst = ws_new.cell(row=new_row, column=new_col)

            value = None
            if src_cell_val.value is not None:
                value = src_cell_val.value
            elif not _is_formula_cell(src_cell_fmt):
                value = src_cell_fmt.value

            # Copy rich text properly
            try:
                if hasattr(src_cell_fmt, "rich_text") and src_cell_fmt.rich_text:
                    rt = CellRichText()
                    for block in src_cell_fmt.rich_text:
                        if isinstance(block, TextBlock):
                            rt.append(deepcopy(block))
                    dst.rich_text = rt
                else:
                    dst.value = value
            except Exception:
                dst.value = value

            try:
                if getattr(src_cell_fmt, "comment", None) is not None:
                    dst.comment = deepcopy(src_cell_fmt.comment)
            except Exception:
                pass

            try:
                if src_cell_fmt.has_style:
                    dst.font = copy(src_cell_fmt.font)
                    dst.border = copy(src_cell_fmt.border)
                    dst.fill = copy(src_cell_fmt.fill)
                    dst.number_format = src_cell_fmt.number_format
                    dst.protection = copy(src_cell_fmt.protection)
                    dst.alignment = copy(src_cell_fmt.alignment)
            except Exception:
                pass

        try:
            rd = ws_fmt.row_dimensions.get(r)
            if rd is not None and getattr(rd, "height", None) is not None:
                ws_new.row_dimensions[new_row].height = rd.height
        except Exception:
            pass

        new_row += 1

    total_rows = new_row - 1

    for src_col, new_col in mapping.items():
        try:
            src_letter = get_column_letter(src_col)
            new_letter = get_column_letter(new_col)
            cd = ws_fmt.column_dimensions.get(src_letter)
            if cd is not None and getattr(cd, "width", None) is not None:
                ws_new.column_dimensions[new_letter].width = cd.width
        except Exception:
            pass

    # Set Summary to open at A1 and be active
    for s in wb_fmt.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws_new.sheet_view.tabSelected = True
    wb_fmt.active = wb_fmt.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Save workbook (this does not modify Group cells' formulas — we only read from Group)
    wb_fmt.save(file_path)
    print(f"Step 5: SUMMARY completed on {file_path}")

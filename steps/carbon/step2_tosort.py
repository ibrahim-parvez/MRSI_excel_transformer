import os
import time
import traceback
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter

def _try_force_excel_recalc(file_path, timeout=5.0):
    """
    Try to open the workbook in Excel via xlwings, calculate, save and close.
    Returns True on success, False on failure (e.g. xlwings not installed or Excel not available).
    """
    try:
        import xlwings as xw
    except Exception:
        return False

    try:
        app = xw.App(visible=False)
        # Give Excel a moment to start
        time.sleep(0.2)
        book = app.books.open(os.path.abspath(file_path))
        # Force a full recalculation
        book.app.calculate()
        # Wait a little for Excel to finish
        time.sleep(min(1.0, timeout))
        book.save()
        book.close()
        app.quit()
        return True
    except Exception:
        try:
            # Best-effort cleanup
            app.quit()
        except Exception:
            pass
        return False


def step2_tosort(file_path, filter_choice="Last 6"):
    """
    Step 2: TO SORT
    Copies rows from 'Data' into 'To Sort' but converts formulas into raw values in To Sort.
    Data sheet keeps its formulas.
    Attempts to force Excel recalc (via xlwings) so cached values exist; if recalculation fails,
    the code will still copy whatever cached values exist (may be None for some formula cells).
    Finally: applies autofilter on column Q and hides rows not matching "last 6".
    """

    source_sheet = "Data"
    new_sheet_name = "To Sort"

    # First: try to force Excel to recalc & save (so workbook will have cached values)
    recalc_ok = _try_force_excel_recalc(file_path)
    if not recalc_ok:
        # Not fatal — we'll continue, but warn the user in the logs (print).
        print("Warning: unable to force Excel recalculation (xlwings missing or failed).")
        print("If Data contains formulas without cached values, To Sort may have empty cells for those formulas.")
        # You can optionally call GUI's refresh routine manually if desired.

    # Load two workbook views:
    # wb (data_only=False) is used as the target workbook we will write the "To Sort" sheet into
    # wb_values (data_only=True) used to read *calculated values* (not formulas)
    wb = load_workbook(file_path, data_only=False)
    wb_values = load_workbook(file_path, data_only=True)

    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in workbook. Run Step 1 first.")
    if source_sheet not in wb_values.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in values workbook. Run Step 1 first.")

    # Remove old To Sort if present (from the formula workbook)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # Source worksheets
    ws_source = wb[source_sheet]            # has formulas preserved
    ws_source_values = wb_values[source_sheet]  # values_only view (cached values)

    # Create To Sort sheet to the LEFT of Data sheet
    ws_new = wb.create_sheet(new_sheet_name, index=wb.index(ws_source))

    # Columns D, E, F = 4,5,6 (1-based)
    text_cols = {4, 5, 6}

    # Copy *values only* from ws_source_values into ws_new
    # Use iter_rows(values_only=True) for robust value extraction.
    max_col_idx = ws_source.max_column
    max_row_idx = ws_source_values.max_row if ws_source_values.max_row > 0 else ws_source.max_row

    for r_idx, row in enumerate(ws_source_values.iter_rows(values_only=True), start=1):
        # row is a tuple of values (length may be <= max_col_idx)
        for c_idx, val in enumerate(row, start=1):
            if c_idx in text_cols and val is not None:
                # Convert to string to trigger Excel green triangle
                val = str(val)
            ws_new.cell(row=r_idx, column=c_idx, value=val)
        # if a row is shorter than max_col_idx, fill remaining columns with None explicitly
        if len(row) < max_col_idx:
            for c_idx in range(len(row) + 1, max_col_idx + 1):
                ws_new.cell(row=r_idx, column=c_idx, value=None)

    # Apply autofilter across full used range (based on source's max row/col)
    last_col_letter = get_column_letter(max_col_idx)
    last_row = max_row_idx
    ws_new.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Apply filter specifically to column Q based on selected option
    filter_choice = (filter_choice or "Last 6").strip().lower()

    target_filter_index = 16  # column Q
    try:
        ws_new.auto_filter.add_filter_column(target_filter_index, [filter_choice])
        ws_new.auto_filter.add_sort_condition(f"Q2:Q{last_row}")
    except Exception:
        pass

    # Hide rows not matching filter (unless "All")
    if filter_choice != "all":
        for r in range(2, last_row + 1):
            val = ws_new.cell(row=r, column=17).value
            if not val or val.lower() != filter_choice:
                ws_new.row_dimensions[r].hidden = True

    # Activate new sheet and set selection
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws_new.sheet_view.tabSelected = True
    wb.active = wb.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Save the workbook (this writes To Sort into the same workbook that still has Data formulas)
    wb.save(file_path)
    print(f"Step 2: TO SORT completed on {file_path}")
    if not recalc_ok:
        print("Note: xlwings recalculation was not run. If To Sort contains blanks in R–AA,")
        print("open the workbook in Excel and save once (or enable auto-calc), then re-run Step 2.")


# End of step2_tosort

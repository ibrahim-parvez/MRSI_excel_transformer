from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection

def step3_last6(file_path):
    """
    Step 3: LAST 6
    Creates a new sheet "Last 6" to the LEFT of 'To Sort' sheet.
    Copies only rows where column Q = "last 6".

    Special rule:
      - Columns labeled 'Comment', 'Identifier 2', and 'Analysis'
        are forced to text (string) to trigger Excel's green flag.
    """

    source_sheet = "To Sort"
    new_sheet_name = "Last 6"

    # Load workbook (values only)
    wb = load_workbook(file_path, data_only=False)
    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found. Run Step 2 first.")

    # Remove old sheet if it exists
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    ws_source = wb[source_sheet]

    # Insert new sheet immediately to the left of To Sort
    ws_new = wb.create_sheet(new_sheet_name, index=wb.index(ws_source))

    # Copy headers (always row 1)
    header_map = {}  # map header names → column indices
    for col_idx, cell in enumerate(ws_source[1], start=1):
        header_val = str(cell.value).strip() if cell.value else ""
        ws_new.cell(row=1, column=col_idx, value=header_val)
        header_map[header_val.lower()] = col_idx

    # Identify special columns for text conversion
    special_headers = {"comment", "identifier 2", "analysis"}
    special_cols = [idx for name, idx in header_map.items() if name in special_headers]

    # Column Q index (1-based)
    col_q = 17

    # Copy rows where Q == "last 6"
    new_row_num = 2
    for r in range(2, ws_source.max_row + 1):
        val_q = ws_source.cell(row=r, column=col_q).value
        if str(val_q).strip().lower() != "last 6":
            continue  # skip rows that are not "last 6"

        row = ws_source[r]
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            if col_idx in special_cols and val is not None:
                ws_new.cell(row=new_row_num, column=col_idx, value=str(val))
            else:
                ws_new.cell(row=new_row_num, column=col_idx, value=val)
        new_row_num += 1

    # Ensure sheet opens at A1 and is active
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws_new.sheet_view.tabSelected = True
    wb.active = wb.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Save workbook
    wb.save(file_path)
    print(f"Step 3: LAST 6 completed on {file_path}")

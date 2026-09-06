import os
import shutil
import re
import time
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.comments import Comment
import tempfile
from copy import copy
from datetime import datetime
import xlwings as xw
import sys
import subprocess

# --- Matplotlib for high-quality graphing ---
import matplotlib
matplotlib.use('Agg')  # Forces headless, thread-safe rendering for PyQt6
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from PyQt6.QtCore import QThread, pyqtSignal
import utils.settings as settings
from utils.common_utils import save_workbook_atomic
from utils.excel_engine import ExcelEngine, set_shared_engine, clear_shared_engine

# ---- Import carbonate modules ----
from processors.carbonate.step1_data import step1_data_carbonate
from processors.carbonate.step2_tosort import step2_tosort_carbonate
from processors.carbonate.step3_last6 import step3_last6_carbonate
from processors.carbonate.step4_pre_group import step4_pre_group_carbonate
from processors.carbonate.step5_group import step5_group_carbonate
from processors.carbonate.step6_normalization import step6_normalization_carbonate
from processors.carbonate.step7_report import step7_report_carbonate

class CarbonateCombineWorker(QThread):
    log = pyqtSignal(str, str)
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    stopped_early = pyqtSignal()

    def __init__(self, params):
        super().__init__()
        self.params = params
        self._is_running = True

    def stop(self):
        self._is_running = False

    def copy_cell_exact(self, src_cell, tgt_cell):
        tgt_cell.value = src_cell.value
        
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.alignment = copy(src_cell.alignment)
            
        if src_cell.comment:
            tgt_cell.comment = Comment(src_cell.comment.text, src_cell.comment.author)

    def get_base_reference_name(self, identifier, ref_settings):
        if not identifier: return None
        raw_text = str(identifier).upper().strip()
        text_clean = re.sub(r'[\s\-_]+', '', raw_text)
        text_no_std = text_clean.replace("STD", "")

        if text_clean.startswith("CO2") or text_clean.startswith("HECO2") or text_clean.startswith("C02") or text_clean.startswith("HEC02"):
            return "HeCO2"

        for std in ref_settings:
            std_name = std.get("col_c")
            if not std_name: continue
            
            std_clean = re.sub(r'[\s\-_]+', '', str(std_name).upper())
            std_no_std = std_clean.replace("STD", "")
            
            if std_clean in text_clean:
                return std_name
            if len(std_no_std) >= 4 and std_no_std in text_no_std:
                return std_name
        return None

    def _is_reference_section_end(self, ws, row, max_col):
        """
        True once `row` marks the end of the reference-material section.

        Step 6 closes that section with a two-row dark grey band and then
        repeats the column headers above the samples. Both have an empty
        Identifier column, so the caller's "stop at the next non-reference
        identifier" rule cannot see them.
        """
        # a) the dark grey divider band Step 6 draws across the sheet
        for c in range(1, min(max_col, 5) + 1):
            rgb = getattr(ws.cell(row=row, column=c).fill.start_color, "rgb", None)
            if not (isinstance(rgb, str) and rgb.endswith("808080")):
                break
        else:
            return True

        # b) the repeated header row: column A holds text ("Row"/"Line") where
        #    a real measurement row holds its run number.
        first = ws.cell(row=row, column=1).value
        return isinstance(first, str) and first.strip() != ""

    def parse_timestamp(self, ts_val):
        if isinstance(ts_val, datetime):
            return ts_val
        if isinstance(ts_val, str):
            try:
                match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})', ts_val)
                if match:
                    return datetime.strptime(match.group(1).replace('-', '/'), "%Y/%m/%d %H:%M:%S")
            except:
                pass
        return datetime.min

    def run(self):
        # Temp dir must always be created now to store the matplotlib generated images
        temp_dir = tempfile.mkdtemp(prefix="mrsi_combine_tmp_")
        engine = None
        master_ref_settings_carb = None
        master_ref_settings_co2 = None
        original_yield = None
        original_co2 = None
        
        try:
            mode = "carbonate"
            files_data = self.params["file_list"]
            output_path = self.params["output_path"]
            protect_originals = self.params.get("protect_originals", False)
            
            # =========================================================================
            # SNAPSHOT MASTER SETTINGS (Don't permanently modify them)
            # =========================================================================
            master_ref_settings_carb = settings.get_setting("REFERENCE_MATERIALS", sub_key="Carbonate") or []
            master_ref_settings_co2 = settings.get_setting("REFERENCE_MATERIALS", sub_key="CO2") or []
            
            original_yield = settings.get_setting("CALC_YIELD")
            original_co2 = settings.get_setting("CALC_CO2")
            
            # =========================================================================
            # TEMPORARILY INJECT UI TOGGLES INTO RUNTIME SETTINGS
            # (Matches exactly what step6_normalization_carbonate.py looks for)
            # =========================================================================
            calc_yield_global = self.params.get("calc_yield", False)
            calc_co2_global = self.params.get("calc_co2", False)
            
            settings.set_setting("CALC_YIELD", calc_yield_global)
            settings.set_setting("CALC_CO2", calc_co2_global)
            
            total_steps = len(files_data) * 8 + 2
            current_step = 0

            self.log.emit(f"Starting Carbonate Process & Combine for {len(files_data)} files...", "white")

            combined_data = {} 
            
            if protect_originals:
                self.log.emit("Working on temporary copies to protect originals.", "white")

            self.log.emit("Starting background Excel engine...", "white")
            engine = ExcelEngine(log=lambda msg: self.log.emit(msg, "white"), keep_alive=True)
            # Publish the engine so the step processors that also need Excel
            # (Step 2 To Sort, Step 7 Report) reuse this instance instead of
            # starting and quitting their own. On Windows, quitting a second
            # instance force-kills this one via xlwings' zombie sweep.
            # See utils/excel_engine.py.
            set_shared_engine(engine)
            engine.app  # start Excel now, so a failure is reported up front

            def local_refresh(file_path):
                # Restarts Excel and retries automatically if the instance dies.
                engine.refresh(file_path, settle=1.0)

            for idx, file_info in enumerate(files_data):
                if not self._is_running: return self.stopped_early.emit()

                raw_file = file_info["path"]
                sheet_name = file_info["sheet"]
                target_file = raw_file
                
                # =========================================================================
                # BUILD TEMPORARY RULES LISTS USING EXISTING HEURISTICS
                # =========================================================================
                file_refs = file_info.get("references", [])
                file_samples = file_info.get("samples", [])
                
                def build_local_refs(master_list):
                    local_list = []
                    # A. Keep master references UNLESS explicitly dragged into Samples
                    for m_ref in master_list:
                        m_name = str(m_ref.get("col_c", "")).strip().upper()
                        is_in_samples = any(str(s).strip().upper() == m_name for s in file_samples)
                        if not is_in_samples:
                            local_list.append(m_ref)
                            
                    # B. Add newly dragged References, skipping variants mapped to a base name
                    for f_ref in file_refs:
                        mapped_name = self.get_base_reference_name(f_ref, local_list)
                        if not mapped_name:
                            local_list.append({
                                "col_c": f_ref,
                                "col_d": "", "col_e": "", "col_f": "", "col_g": "", "col_h": "",
                                "color": "black"
                            })
                    return local_list

                # Temporarily apply local references for Carbonate
                local_ref_settings_carb = build_local_refs(master_ref_settings_carb)
                settings.set_setting("REFERENCE_MATERIALS", local_ref_settings_carb, sub_key="Carbonate")
                
                # Temporarily apply local references for CO2 (if enabled)
                if calc_co2_global:
                    local_ref_settings_co2 = build_local_refs(master_ref_settings_co2)
                    settings.set_setting("REFERENCE_MATERIALS", local_ref_settings_co2, sub_key="CO2")

                ref_settings = local_ref_settings_carb  # Keep for charting later
                # =========================================================================

                if protect_originals:
                    filename = os.path.basename(raw_file)
                    target_file = os.path.join(temp_dir, f"{idx}_{filename}")
                    shutil.copy(raw_file, target_file)

                self.log.emit("-" * 40, "white")
                self.log.emit(f"⚙️ Processing File {idx+1}/{len(files_data)}: {os.path.basename(raw_file)}", "white")
                
                step2_carbonate = lambda: step2_tosort_carbonate(target_file, "")
                
                # Original step processors without modification
                step_order = [
                    ("Step 1: Data", lambda: step1_data_carbonate(target_file, sheet_name)),
                    ("Step 2: To Sort", lambda: (local_refresh(target_file), step2_carbonate())), 
                    ("Step 3: Last 6", lambda: step3_last6_carbonate(target_file)),
                    ("Step 4: Pre-Group", lambda: step4_pre_group_carbonate(target_file)),
                    ("Step 5: Group", lambda: step5_group_carbonate(target_file)),
                    ("Step 6: Normalization", lambda: step6_normalization_carbonate(target_file)),
                    ("Step 7: Report", lambda: step7_report_carbonate(target_file)),
                ]

                for name, func in step_order:
                    if not self._is_running: return self.stopped_early.emit()
                    self.log.emit(f"▶  Running {name}...", "white")
                    time.sleep(1.0)
                    try:
                        func()
                        self.log.emit(f"✔  {name} Completed", "green")
                    except Exception as e:
                        raise Exception(f"{name} Failed on file {os.path.basename(target_file)}: {str(e)}")
                        
                    current_step += 1
                    self.progress.emit(current_step, total_steps, f"{os.path.basename(target_file)} - {name}")

                self.log.emit(f"🔄 Preparing final calculations for {os.path.basename(target_file)}...", "white")
                local_refresh(target_file)
                time.sleep(1.5)

                self.log.emit(f"Extracting standards from {os.path.basename(target_file)}...", "white")
                
                wb = openpyxl.load_workbook(target_file, data_only=True)
                if "Normalization_DNT" not in wb.sheetnames:
                    raise Exception(f"Sheet 'Normalization_DNT' was not created in {os.path.basename(target_file)}")
                    
                ws = wb["Normalization_DNT"]
                
                data_header_row = 1
                for r in range(1, ws.max_row + 1):
                    val = str(ws.cell(row=r, column=3).value or "").strip().lower()
                    if "identifier" in val or "time code" in str(ws.cell(row=r, column=2).value or "").strip().lower():
                        data_header_row = r
                        break

                # --- Dynamic Column Scanning ---
                max_data_col = 17 
                for r in range(1, max(2, data_header_row)):
                    for c in range(100, 17, -1):
                        if ws.cell(row=r, column=c).value is not None:
                            if c > max_data_col:
                                max_data_col = c
                            break
                
                max_col_to_copy = max_data_col + 1
                
                # --- CARBONATE EXTRACTION LOGIC ---
                end_blue_box = max(15, data_header_row - 1)
                file_blue_box = []
                for r in range(1, end_blue_box + 1):
                    row_cells = [ws.cell(row=r, column=c) for c in range(1, 18)] 
                    file_blue_box.append(row_cells)
                    
                file_data_header = [ws.cell(row=data_header_row, column=c) for c in range(1, max_col_to_copy + 1)]
                
                file_blocks = {}
                current_mat = None
                recording = False
                
                for r in range(data_header_row + 1, ws.max_row + 1):
                    cell_c = ws.cell(row=r, column=3).value
                    cell_b = ws.cell(row=r, column=2).value 
                    
                    if cell_c is not None and str(cell_c).strip() != "":
                        base_mat = self.get_base_reference_name(cell_c, ref_settings)
                        if base_mat:
                            current_mat = base_mat
                            recording = True 
                            
                            if current_mat not in file_blocks:
                                file_blocks[current_mat] = {
                                    'filename': os.path.basename(target_file),
                                    'timestamp': None,
                                    'rows': []
                                }
                                
                            if file_blocks[current_mat]['timestamp'] is None and cell_b:
                                ts = self.parse_timestamp(cell_b)
                                if ts != datetime.min:
                                    file_blocks[current_mat]['timestamp'] = ts
                        else:
                            recording = False
                    elif recording and self._is_reference_section_end(ws, r, max_col_to_copy):
                        # Blank Identifier, but the standard's block has ended.
                        recording = False

                    if recording:
                        row_cells = [ws.cell(row=r, column=c) for c in range(1, max_col_to_copy + 1)] 
                        file_blocks[current_mat]['rows'].append(row_cells)
                        
                # Trim trailing blank rows
                for mat, block_data in file_blocks.items():
                    rows = block_data['rows']
                    while rows:
                        is_empty = True
                        
                        for cell in rows[-1][:17]:
                            if cell.value is not None and str(cell.value).strip() != "":
                                is_empty = False
                                break
                        
                        if is_empty:
                            rows.pop() 
                        else:
                            break
                            
                for mat, block_data in file_blocks.items():
                    if mat not in combined_data:
                        combined_data[mat] = []
                    combined_data[mat].append({
                        'filename': block_data['filename'],
                        'timestamp': block_data['timestamp'],
                        'blue_box': file_blue_box,
                        'data_header': file_data_header,
                        'block_rows': block_data['rows']
                    })

                wb.close()
                current_step += 1
                self.progress.emit(current_step, total_steps, f"Extracted {os.path.basename(target_file)}")

            if not self._is_running: return self.stopped_early.emit()
            self.log.emit("=" * 40, "white")
            self.log.emit(f"Compiling {len(combined_data)} Standard sheets...", "white")
            
            append_mode = self.params.get("append_mode", False)
            
            if append_mode and os.path.exists(output_path):
                self.log.emit(f"Loading existing file to append and sort data: {os.path.basename(output_path)}", "white")
                out_wb = openpyxl.load_workbook(output_path)
            else:
                out_wb = openpyxl.Workbook()
                out_wb.remove(out_wb.active) 
            
            grey_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            black_font = openpyxl.styles.Font(bold=True, color="000000")
            
            color_hex_map = {
                "red": "FF0000", "blue": "0000FF", "darkblue": "00008B", 
                "orange": "E46C0A", "green": "00B050", "lightblue": "5DADE2", 
                "black": "000000"
            }

            # =======================================================
            # COMPILATION: CHRONOLOGICAL INJECTION & GRAPH REBUILDING
            # =======================================================
            for mat_name, files_data_list in combined_data.items():
                clean_title = str(mat_name)[:31] 
                
                is_new_sheet = False
                if clean_title in out_wb.sheetnames:
                    ws_out = out_wb[clean_title]
                else:
                    ws_out = out_wb.create_sheet(title=clean_title)
                    is_new_sheet = True
                    ws_out.freeze_panes = "A2"
                
                files_data_list.sort(key=lambda x: x['timestamp'] if x['timestamp'] else datetime.min)
                
                mat_color = "000000"
                for std in ref_settings:
                    if std.get("col_c") == mat_name:
                        c_name = std.get("color", "black").lower()
                        mat_color = color_hex_map.get(c_name, "000000")
                        break
                
                if is_new_sheet and files_data_list and files_data_list[0]['data_header']:
                    for c_idx, src_cell in enumerate(files_data_list[0]['data_header']):
                        tgt_cell = ws_out.cell(row=1, column=1 + c_idx)
                        self.copy_cell_exact(src_cell, tgt_cell)
                    
                    # --- Add "Normalized VPDB" label in column R (18) ---
                    norm_label_cell = ws_out.cell(row=1, column=18, value="Normalized VPDB")
                    norm_label_cell.font = openpyxl.styles.Font(bold=True, color="000000")
                    norm_label_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    
                    # ---Add "VSMOW" label in column AA (27) ---
                    vsmow_label_cell = ws_out.cell(row=1, column=27, value="VSMOW")
                    vsmow_label_cell.font = openpyxl.styles.Font(bold=True, color="000000")
                    vsmow_label_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                
                divider_width = 17 # Fixed width for the Data From row (A to Q)
                
                # BLOCK 1: WRITING & INJECTING DATA CHRONOLOGICALLY
                if is_new_sheet:
                    current_out_row = 2
                    for file_data in files_data_list:
                        current_width = len(file_data['data_header'])
                        offset_col = current_width + 2
                        
                        ws_out.merge_cells(start_row=current_out_row, start_column=1, end_row=current_out_row, end_column=divider_width)
                        div_cell = ws_out.cell(row=current_out_row, column=1, value=f"Data from: {file_data['filename']}")
                        div_cell.fill = grey_fill
                        div_cell.font = black_font
                        div_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                        for c in range(1, divider_width + 1):
                            ws_out.cell(row=current_out_row, column=c).fill = grey_fill
                            
                        current_out_row += 1
                        start_data_row = current_out_row
                        
                        for r_idx, row_cells in enumerate(file_data['block_rows']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=1 + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        for r_idx, row_cells in enumerate(file_data['blue_box']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=offset_col + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        end_data_row = start_data_row + len(file_data['block_rows']) - 1
                        
                        max_written = max(end_data_row, start_data_row + len(file_data['blue_box']) - 1)
                        current_out_row = max_written + 3 
                        
                    for c in range(1, offset_col + current_width):
                        ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15

                else:
                    # Append Mode: Dynamically inject blocks in chronological order
                    ws_out._charts = [] # Clear old charts to prevent corruption
                    
                    existing_blocks = []
                    for r in range(2, ws_out.max_row + 1):
                        val = ws_out.cell(row=r, column=1).value
                        if isinstance(val, str) and val.startswith("Data from:"):
                            ts_val = ws_out.cell(row=r+1, column=2).value
                            ts = self.parse_timestamp(ts_val)
                            existing_blocks.append({'start_row': r, 'timestamp': ts})
                    
                    last_real_row = 2
                    for r in range(ws_out.max_row, 1, -1):
                        v1 = ws_out.cell(row=r, column=1).value
                        v2 = ws_out.cell(row=r, column=2).value
                        if (v1 and str(v1).strip() != "") or (v2 and str(v2).strip() != ""):
                            last_real_row = r
                            break
                    
                    for file_data in files_data_list:
                        new_ts = file_data['timestamp'] or datetime.min
                        
                        insert_idx = last_real_row + 3
                        
                        for eb in existing_blocks:
                            if eb['timestamp'] > new_ts:
                                insert_idx = eb['start_row']
                                break
                        
                        block_height = len(file_data['block_rows'])
                        blue_box_height = len(file_data['blue_box'])
                        total_height_needed = max(block_height, blue_box_height) + 3
                        
                        if insert_idx <= last_real_row:
                            ws_out.insert_rows(insert_idx, amount=total_height_needed)
                            for eb in existing_blocks:
                                if eb['start_row'] >= insert_idx:
                                    eb['start_row'] += total_height_needed
                            last_real_row += total_height_needed
                        else:
                            last_real_row += total_height_needed
                        
                        current_width = len(file_data['data_header'])
                        offset_col = current_width + 2
                        
                        div_cell = ws_out.cell(row=insert_idx, column=1, value=f"Data from: {file_data['filename']}")
                        div_cell.fill = grey_fill
                        div_cell.font = black_font
                        div_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                        for c in range(1, divider_width + 1):
                            ws_out.cell(row=insert_idx, column=c).fill = grey_fill
                            
                        start_data_row = insert_idx + 1
                        
                        for r_idx, row_cells in enumerate(file_data['block_rows']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=1 + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        for r_idx, row_cells in enumerate(file_data['blue_box']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=offset_col + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)

                # BLOCK 1.5: CLEAN MERGING SWEEP (Fixes OpenPyXL format bugs)
                # Unmerge any remaining corrupted cells to prevent errors
                for m_range in list(ws_out.merged_cells.ranges):
                    try:
                        ws_out.unmerge_cells(str(m_range))
                    except Exception:
                        pass
                
                # Re-apply the merge and grey styling to every single header row perfectly
                for r in range(2, ws_out.max_row + 1):
                    val = ws_out.cell(row=r, column=1).value
                    if isinstance(val, str) and val.startswith("Data from:"):
                        ws_out.merge_cells(start_row=r, start_column=1, end_row=r, end_column=divider_width)
                        for c in range(1, divider_width + 1):
                            cell = ws_out.cell(row=r, column=c)
                            cell.fill = grey_fill
                            if c == 1:
                                cell.font = black_font
                                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

                # =======================================================
                # BLOCK 2: UNIFIED CHARTS (EXCEL NATIVE + MATPLOTLIB)
                # =======================================================
                real_max_row = 1
                for r in range(ws_out.max_row, 0, -1):
                    row_has_data = False
                    for c in range(1, 40):
                        val = ws_out.cell(row=r, column=c).value
                        if val is not None and str(val).strip() != "":
                            row_has_data = True
                            break
                    if row_has_data:
                        real_max_row = r
                        break

                # Ensure Time Codes in Column B are actual datetime objects
                for r in range(2, real_max_row + 1):
                    val_b = ws_out.cell(row=r, column=2).value
                    if isinstance(val_b, str):
                        try:
                            match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})', val_b)
                            if match:
                                dt_val = datetime.strptime(match.group(1).replace('-', '/'), "%Y/%m/%d %H:%M:%S")
                                ws_out.cell(row=r, column=2).value = dt_val
                                ws_out.cell(row=r, column=2).number_format = 'yyyy/mm/dd'
                        except:
                            pass

                all_data_chunks = []
                current_start = None
                
                for r in range(2, real_max_row + 1):
                    val_b = ws_out.cell(row=r, column=2).value
                    is_valid_time = isinstance(val_b, datetime) or (isinstance(val_b, str) and re.search(r'\d{4}[-/]\d{2}', str(val_b)))
                    
                    if is_valid_time:
                        if current_start is None:
                            current_start = r
                    else:
                        if current_start is not None:
                            all_data_chunks.append((current_start, r - 1))
                            current_start = None
                            
                if current_start is not None:
                    all_data_chunks.append((current_start, real_max_row))

                raw_columns = [
                    (11, "δ¹³C RAW"),
                    (14, "δ¹⁸O RAW")
                ]
                
                norm_columns = []
                max_norm_col = len(files_data_list[0]['data_header'])
                
                for i in range(19, max_norm_col + 1):
                    has_data_in_col = False
                    # Check a few rows down to ensure the column actually contains data values
                    for r in range(3, 15):
                        if ws_out.cell(row=r, column=i).value is not None:
                            has_data_in_col = True
                            break
                            
                    if has_data_in_col:
                        val = ws_out.cell(row=1, column=i).value
                        if val and str(val).strip():
                            hdr_text = str(val).strip()
                            if hdr_text.upper() not in ["STDEV", "N"]:
                                norm_columns.append((i, f"Normalized ({hdr_text})"))
                        else:
                            # If header is blank but it has data, give it a default name
                            norm_columns.append((i, f"Normalized (Col {openpyxl.utils.get_column_letter(i)})"))

                # PART A: EXCEL INTERACTIVE CHARTS
                chart_start_row = real_max_row + 2
                chart_height = 14 
                chart_width = 22 
                col_spacing = 12  
                row_spacing = 30  # Increased to prevent overlap
                
                excel_charts_to_plot = []
                
                # Top Row: RAW charts side-by-side
                for idx, (col_idx, title) in enumerate(raw_columns):
                    col_pos = 3 + (idx * col_spacing)
                    excel_charts_to_plot.append((col_idx, title, chart_start_row, col_pos))
                    
                # Lower Rows: Normalized charts stacked in 2-column grid
                norm_start_row = chart_start_row + row_spacing
                for idx, (col_idx, title) in enumerate(norm_columns):
                    row_pos = norm_start_row + (idx // 2) * row_spacing
                    col_pos = 3 + (idx % 2) * col_spacing
                    excel_charts_to_plot.append((col_idx, title, row_pos, col_pos))

                max_excel_chart_row = chart_start_row

                for col_idx, chart_title, c_row, c_col in excel_charts_to_plot:
                    chart = ScatterChart()
                    chart.title = chart_title
                    
                    chart.x_axis.title = "Date"
                    chart.y_axis.title = chart_title
                    chart.legend = None 
                    chart.height = chart_height
                    chart.width = chart_width
                    
                    chart.x_axis.majorGridlines = ChartLines()
                    chart.y_axis.majorGridlines = ChartLines()
                    
                    chart.x_axis.tickLblPos = "nextTo"
                    chart.y_axis.tickLblPos = "nextTo"
                    chart.x_axis.majorTickMark = "out"
                    chart.y_axis.majorTickMark = "out"
                    
                    chart.y_axis.numFmt = "0.00"
                    chart.y_axis.numFmtLinked = False
                    chart.x_axis.number_format = 'yyyy/mm/dd'
                    chart.x_axis.numFmtLinked = False
                    
                    y_values_list = []
                    has_data = False
                    
                    for chunk in all_data_chunks:
                        s_row, e_row = chunk
                        valid_ranges = []
                        c_start = None
                        
                        for r in range(s_row, e_row + 1):
                            val_y = ws_out.cell(row=r, column=col_idx).value
                            if isinstance(val_y, (int, float)):
                                y_values_list.append(val_y)
                                if c_start is None:
                                    c_start = r
                            else:
                                if c_start is not None:
                                    valid_ranges.append((c_start, r - 1))
                                    c_start = None
                        
                        if c_start is not None:
                            valid_ranges.append((c_start, e_row))
                            
                        for sub_s, sub_e in valid_ranges:
                            xvalues = Reference(ws_out, min_col=2, min_row=sub_s, max_row=sub_e)
                            yvalues = Reference(ws_out, min_col=col_idx, min_row=sub_s, max_row=sub_e)
                            
                            series = Series(values=yvalues, xvalues=xvalues, title_from_data=False)
                            
                            series.graphicalProperties.line.noFill = True
                            
                            series.marker.symbol = "circle"
                            series.marker.size = 5
                            series.marker.graphicalProperties.solidFill = mat_color
                            series.marker.graphicalProperties.line.solidFill = mat_color
                            
                            chart.series.append(series)
                            has_data = True
                            
                    # Calculate Y-Axis padding (Generous 40% margin)
                    if y_values_list:
                        min_y = min(y_values_list)
                        max_y = max(y_values_list)
                        y_range = max_y - min_y
                        
                        if y_range == 0:
                            padding = abs(min_y) * 0.1 if min_y != 0 else 1.0
                        else:
                            padding = y_range * 0.40  # 40% margin above and below
                            
                        chart.y_axis.scaling.min = round(min_y - padding, 2)
                        chart.y_axis.scaling.max = round(max_y + padding, 2)
                            
                    if has_data:
                        col_letter = openpyxl.utils.get_column_letter(c_col)
                        ws_out.add_chart(chart, f"{col_letter}{c_row}")
                        if c_row > max_excel_chart_row:
                            max_excel_chart_row = c_row

                # PART B: MATPLOTLIB STATIC IMAGE CHARTS
                mpl_start_row = max_excel_chart_row + row_spacing + 5
                mpl_col_spacing = 12  
                mpl_row_spacing = 32  # Increased to prevent overlap
                
                mpl_charts_to_plot = []
                
                for idx, (col_idx, title) in enumerate(raw_columns):
                    col_pos = 3 + (idx * mpl_col_spacing)
                    mpl_charts_to_plot.append((col_idx, title, mpl_start_row, col_pos))
                    
                mpl_norm_start_row = mpl_start_row + mpl_row_spacing
                for idx, (col_idx, title) in enumerate(norm_columns):
                    row_pos = mpl_norm_start_row + (idx // 2) * mpl_row_spacing
                    col_pos = 3 + (idx % 2) * mpl_col_spacing
                    mpl_charts_to_plot.append((col_idx, title, row_pos, col_pos))

                max_mpl_chart_row = mpl_start_row
                hex_color = f"#{mat_color}" if not mat_color.startswith("#") else mat_color

                for col_idx, chart_title, c_row, c_col in mpl_charts_to_plot:
                    x_data = []
                    y_data = []
                    
                    for chunk in all_data_chunks:
                        s_row, e_row = chunk
                        for r in range(s_row, e_row + 1):
                            val_y = ws_out.cell(row=r, column=col_idx).value
                            val_x = ws_out.cell(row=r, column=2).value
                            
                            if isinstance(val_y, (int, float)) and isinstance(val_x, datetime):
                                x_data.append(val_x)
                                y_data.append(val_y)
                                
                    if x_data and y_data:
                        fig, ax = plt.subplots(figsize=(9, 5))
                        ax.scatter(x_data, y_data, color=hex_color, s=40, zorder=3)
                        
                        ax.set_title(f"{chart_title}", fontsize=12, fontweight='bold')
                        ax.set_xlabel("Date", fontsize=10, fontweight='bold')
                        ax.set_ylabel(chart_title, fontsize=10, fontweight='bold')
                        
                        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m/%d'))
                        fig.autofmt_xdate(rotation=45)
                        
                        ax.grid(True, linestyle='--', alpha=0.7, zorder=0)
                        
                        # Generous 40% Padding for Matplotlib Y-Axis
                        min_y, max_y = min(y_data), max(y_data)
                        y_range = max_y - min_y
                        pad = y_range * 0.40 if y_range != 0 else abs(min_y) * 0.1
                        ax.set_ylim(min_y - pad, max_y + pad)
                        
                        fig.tight_layout()
                        
                        img_filename = f"chart_img_{mat_name}_{col_idx}.png"
                        safe_filename = "".join([c for c in img_filename if c.isalpha() or c.isdigit() or c in (' ', '.', '_')]).rstrip()
                        img_path = os.path.join(temp_dir, safe_filename)
                        
                        fig.savefig(img_path, dpi=100)
                        plt.close(fig) 
                        
                        col_letter = openpyxl.utils.get_column_letter(c_col)
                        img = ExcelImage(img_path)
                        ws_out.add_image(img, f"{col_letter}{c_row}")
                        
                        if c_row > max_mpl_chart_row:
                            max_mpl_chart_row = c_row

                ws_out.cell(row=max_mpl_chart_row + mpl_row_spacing + 5, column=1).value = " "

            save_workbook_atomic(out_wb, output_path)
            self.log.emit("-" * 50, "white")
            self.log.emit(f"✅ Combine Complete! Saved to: {output_path}", "green")
            
            if self.params.get("open_on_complete") and os.path.exists(output_path):
                self.log.emit("Opening combined file...", "white")
                try:
                    if sys.platform == "win32":
                        os.startfile(output_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", output_path])
                    else:
                        subprocess.call(["xdg-open", output_path])
                except Exception as e:
                    self.log.emit(f"Warning: Could not automatically open file: {e}", "white")

            self.progress.emit(total_steps, total_steps, "Done")
            self.finished.emit()

        except Exception as e:
            self.log.emit(f"❌ Critical Error in Combine: {str(e)}", "red")
            self.error.emit(str(e))
            
        finally:
            # =========================================================================
            # RESTORE GLOBAL SETTINGS ONCE DONE
            # =========================================================================
            if master_ref_settings_carb is not None:
                settings.set_setting("REFERENCE_MATERIALS", master_ref_settings_carb, sub_key="Carbonate")
            if master_ref_settings_co2 is not None:
                settings.set_setting("REFERENCE_MATERIALS", master_ref_settings_co2, sub_key="CO2")
                
            if 'original_yield' in locals() and original_yield is not None:
                settings.set_setting("CALC_YIELD", original_yield)
            if 'original_co2' in locals() and original_co2 is not None:
                settings.set_setting("CALC_CO2", original_co2)
            # =========================================================================

            if engine is not None:
                clear_shared_engine(engine)
                engine.shutdown()
                    
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    self.log.emit("Cleaned up temporary workspace.", "white")
                except Exception as e:
                    self.log.emit(f"Warning: Could not completely delete temp directory: {e}", "white")
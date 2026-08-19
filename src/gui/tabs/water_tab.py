import openpyxl
import re
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QCheckBox, QLineEdit, 
    QComboBox, QGroupBox, QFrame, QLabel, QPushButton, QListWidget, QAbstractItemView, QMessageBox, QSizePolicy
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
import utils.settings as settings

# =========================================================================
# BACKGROUND WORKER FOR FETCHING SHEET NAMES
# =========================================================================
class SheetFetcherWorker(QThread):
    """Reads the sheet names of an Excel file in the background."""
    result_ready = pyqtSignal(str, str)  # Emits (file_path, rightmost_sheet_name)

    def __init__(self, file_path, default_fallback):
        super().__init__()
        self.file_path = file_path
        self.default_fallback = default_fallback

    def run(self):
        import pandas as pd 
        sheet_name = self.default_fallback
        
        try:
            xl = pd.ExcelFile(self.file_path)
            if xl.sheet_names:
                sheet_name = xl.sheet_names[-1]
        except Exception:
            pass 
        
        self.result_ready.emit(self.file_path, sheet_name)

# =========================================================================
# MAIN WATER TAB
# =========================================================================
class WaterTab(QWidget):
    references_updated = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.step_cbs = []
        self.select_all_cb = None
        self.sheet_entry = None
        self.filter_combo = None 
        self.step3_label = None 
        self.current_file_path = None
        self.fetcher_thread = None
        self._is_fetching = False
        
        self.base_step_names = [
            "Step 1: Data",
            "Step 2: To Sort",
            "Step 3: Last 6",
            "Step 4: Pre-Group",
            "Step 5: Group",
            "Step 6: Normalization",
            "Step 7: Report"
        ]
        
        # Main layout is Horizontal to hold Left (Steps) and Right (Settings)
        self.layout = QHBoxLayout(self) 
        self.layout.setAlignment(Qt.AlignmentFlag.AlignTop) 
        
        self._create_step_ui()
        self._create_settings_ui()
        self.refresh_step_labels() 

    def _add_divider(self, layout):
        divider = QFrame()
        divider.setFrameShape(QFrame.Shape.HLine)
        divider.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(divider)

    def _create_step_ui(self):
        """Generates the steps UI on the left side."""
        steps_group = QGroupBox("Water Steps")
        steps_group.setFixedWidth(280) # Narrowed box
        steps_layout = QVBoxLayout()
        steps_group.setLayout(steps_layout)

        # Top: Select all toggle
        top_row = QHBoxLayout()
        top_row.addSpacing(10)
        self.select_all_cb = QCheckBox("Select All")
        
        self.select_all_cb.toggled.connect(self.toggle_select_all)
        
        top_row.addWidget(self.select_all_cb)
        top_row.addStretch()
        steps_layout.addLayout(top_row)
        steps_layout.addSpacing(8)
        
        for i, step_name in enumerate(self.base_step_names):
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)
            
            if "Step 3" in step_name:
                cb = QCheckBox() 
                self.step_cbs.append(cb)
                row.addWidget(cb)
                
                self.step3_label = QLabel(step_name)
                self.step3_label.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
                row.addWidget(self.step3_label)
                
                def toggle_cb(event, checkbox=cb):
                    checkbox.setChecked(not checkbox.isChecked())
                self.step3_label.mousePressEvent = toggle_cb
            else:
                cb = QCheckBox(step_name)
                self.step_cbs.append(cb)
                row.addWidget(cb)
            
            cb.toggled.connect(self.update_select_all_state)
            
            row.addStretch()
            steps_layout.addLayout(row)
            
            if i < len(self.base_step_names) - 1:
                self._add_divider(steps_layout)

        steps_layout.addStretch()

        self.layout.addWidget(steps_group)

    def _create_settings_ui(self):
        """Generates the basic settings UI on the right side."""
        settings_group = QGroupBox("Basic Settings")
        settings_layout = QVBoxLayout()
        settings_group.setLayout(settings_layout)
        
        # 1. Sheet Name Entry
        sheet_row = QHBoxLayout()
        self.sheet_entry = QLineEdit("") 
        self.sheet_entry.setPlaceholderText("Sheet name")
        sheet_row.addWidget(QLabel("<b>Initial Sheet Name:</b>"))
        sheet_row.addWidget(self.sheet_entry)
        settings_layout.addLayout(sheet_row)

        # 2. To Sort Filter
        filter_row = QHBoxLayout()
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Last 6", "Last 6 Outliers Excl.", "all", "Amp 44", "delta", "end 11", "ref avg", "sparkline", "start 6"]) 
        self.filter_combo.setCurrentText("Last 6")
        filter_row.addWidget(QLabel("<b>To Sort Filter:</b>"))
        filter_row.addWidget(self.filter_combo)
        settings_layout.addLayout(filter_row)

        # --- Freeze Pane Checkbox ---
        freeze_row = QHBoxLayout()
        self.freeze_pane_cb = QCheckBox("Enable Freeze Pane in Normalization")
        self.freeze_pane_cb.setChecked(True) # Defaults to enabled
        self.freeze_pane_cb.setCursor(Qt.CursorShape.PointingHandCursor)
        freeze_row.addWidget(self.freeze_pane_cb)
        settings_layout.addLayout(freeze_row)
        
        self._add_divider(settings_layout)
        
        # 3. Compact Drag & Drop Sorter
        lists_layout = QHBoxLayout()
        
        # References List
        ref_layout = QVBoxLayout()
        ref_header = QHBoxLayout()
        ref_header.addWidget(QLabel("<b>Reference Materials</b>"))
        ref_header.addStretch()

        # Add an invisible dummy button to balance the header heights ---
        dummy_btn = QPushButton("Placeholder")
        dummy_btn.setStyleSheet("padding: 2px 8px;") # Match the Refresh button exactly
        size_policy = dummy_btn.sizePolicy()
        size_policy.setRetainSizeWhenHidden(True)
        dummy_btn.setSizePolicy(size_policy)
        dummy_btn.hide()
        ref_header.addWidget(dummy_btn)


        ref_layout.addLayout(ref_header)
        
        self.ref_list = QListWidget()
        self.ref_list.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.ref_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.ref_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.ref_list.model().rowsInserted.connect(self.sync_references_to_settings)
        ref_layout.addWidget(self.ref_list)

        # Samples List
        sample_layout = QVBoxLayout()
        sample_header = QHBoxLayout()
        sample_header.addWidget(QLabel("<b>Samples</b>"))
        sample_header.addStretch()
        
        # Compact Auto-Fetch Button
        self.fetch_btn = QPushButton("↻ Refresh")
        self.fetch_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fetch_btn.setToolTip("Fetch unique materials from file")
        self.fetch_btn.setStyleSheet("padding: 2px 8px;")
        self.fetch_btn.clicked.connect(self.fetch_unique_materials)
        sample_header.addWidget(self.fetch_btn)
        
        sample_layout.addLayout(sample_header)
        
        self.sample_list = QListWidget()
        self.sample_list.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.sample_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.sample_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        sample_layout.addWidget(self.sample_list)

        self.ref_list.model().rowsInserted.connect(self.sync_references_to_settings)
        self.ref_list.model().rowsRemoved.connect(self.sync_references_to_settings)
        self.sample_list.model().rowsInserted.connect(self.sync_references_to_settings)
        self.sample_list.model().rowsRemoved.connect(self.sync_references_to_settings)
        
        lists_layout.addLayout(ref_layout)
        lists_layout.addLayout(sample_layout)
        
        settings_layout.addLayout(lists_layout)
        self.layout.addWidget(settings_group)

    def scan_file(self, file_path):
        """Called automatically when a file is dropped in DataToolApp."""
        self.current_file_path = file_path
        
        # Start the background worker to fetch the sheet name securely
        self.fetcher_thread = SheetFetcherWorker(file_path, self.sheet_entry.text())
        self.fetcher_thread.result_ready.connect(self._on_sheet_fetched)
        self.fetcher_thread.start()

    def _on_sheet_fetched(self, file_path, sheet_name):
        """Slot triggered when the SheetFetcherWorker completes."""
        self.sheet_entry.setText(sheet_name)
        self.fetch_unique_materials()

    def fetch_unique_materials(self):
        """Scans the designated sheet for unique materials in the Identifier column."""
        self._is_fetching = True
        
        if not self.current_file_path:
            QMessageBox.warning(self, "No File", "Please select or drop a file first.")
            return
            
        sheet_name = self.sheet_entry.text().strip()
        
        try:
            wb = openpyxl.load_workbook(self.current_file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                QMessageBox.warning(self, "Sheet Error", f"Sheet '{sheet_name}' not found in file.")
                wb.close()
                return
                
            ws = wb[sheet_name]
            
            # Find the "Identifier 1" column (usually C)
            id_col_idx = 3 
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value and str(cell.value).strip().lower() == "identifier 1":
                    id_col_idx = col_idx
                    break
                    
            unique_materials = set()
            for row in ws.iter_rows(min_row=2, min_col=id_col_idx, max_col=id_col_idx, values_only=True):
                val = row[0]
                if val:
                    # Strip run numbers to get the base material name
                    base = re.sub(r"\s*r\d+(\.\d+)?$", "", str(val), flags=re.IGNORECASE).strip()
                    unique_materials.add(base)
            
            wb.close()
            
            # Clear current lists
            self.ref_list.clear()
            self.sample_list.clear()
            
            # Auto-sort logic based on custom Regex Heuristics ---
            def is_heco2(k): 
                return bool(re.search(r'(?i)\b(heco2|co2)\b', str(k))) if k else False
                
            def is_ref(k):
                pat = [r'\bMRSI\b', r'\bMRSI[- ]?\d+\b', r'\bMRSI[- ]?STD', r'\bUSGS']
                return any(re.search(p, str(k).upper()) for p in pat) if k else False
            
            for mat in sorted(unique_materials):
                if is_ref(mat) or is_heco2(mat):
                    self.ref_list.addItem(mat)
                else:
                    self.sample_list.addItem(mat)
                    
        except Exception as e:
            QMessageBox.critical(self, "Fetch Error", f"Failed to fetch materials:\n{e}")
        
        self._is_fetching = False
        self.sync_references_to_settings()
    
    def refresh_step_labels(self):
        calc_mode_step3 = settings.get_setting("CALC_MODE_STEP3")
        
        if self.step3_label:
            if calc_mode_step3 == "Last 6 Outliers Excl.":
                self.step3_label.setText("Step 3: Last 6 <span style='color:red;'><b>Outliers Excluded</b></span>")
            else:
                self.step3_label.setText("Step 3: Last 6")

    # --- UI Logic Methods ---
    def toggle_select_all(self, state):
        is_checked = self.select_all_cb.isChecked()
        for cb in self.step_cbs:
            cb.blockSignals(True)
            cb.setChecked(is_checked)
            cb.blockSignals(False)

    def update_select_all_state(self):
        all_checked = all(cb.isChecked() for cb in self.step_cbs) if self.step_cbs else False
        self.select_all_cb.blockSignals(True)
        self.select_all_cb.setChecked(all_checked)
        self.select_all_cb.blockSignals(False)

    def get_run_parameters(self):
        """Returns the steps dictionary, sheet name, and filter option."""
        steps = {name: cb.isChecked() for name, cb in zip(self.base_step_names, self.step_cbs)}
        sheet_name = self.sheet_entry.text().strip() if self.sheet_entry else "Default_Sheet"
        filter_opt = self.filter_combo.currentText() if self.filter_combo else "N/A"
        return steps, sheet_name, filter_opt

    def get_advanced_parameters(self):
        """Returns the specific settings configured in this tab."""
        refs = [self.ref_list.item(i).text() for i in range(self.ref_list.count())]
        samples = [self.sample_list.item(i).text() for i in range(self.sample_list.count())]

        settings.set_setting("ENABLE_FREEZE_PANE", self.freeze_pane_cb.isChecked())
        
        # Water tab doesn't have Yield or CO2 toggles, so we don't supply them
        return {
            "references": refs,
            "samples": samples
        }

    def _is_mapped_to_existing(self, name_str, existing_mats):
        """Heuristic check to see if a name already exists in the settings table."""
        clean_name = re.sub(r'[\s\-_]+', '', str(name_str).upper())
        no_std = clean_name.replace("STD", "")
        
        for mat in existing_mats:
            std_name = mat.get("col_c", "")
            if not std_name: continue
            std_clean = re.sub(r'[\s\-_]+', '', str(std_name).upper())
            std_no_std = std_clean.replace("STD", "")
            
            if std_clean and std_clean in clean_name: return True
            if std_no_std and len(std_no_std) >= 4 and std_no_std in no_std: return True
        return False

    def sync_references_to_settings(self, *args):
        # Add a 50ms delay to allow the drag-and-drop text to fully populate
        QTimer.singleShot(50, self._do_sync_references)

    def _do_sync_references(self):
        if getattr(self, '_is_fetching', False): return
        if not self.current_file_path: return # Don't wipe tables when no file is loaded
        
        refs_in_ui = [self.ref_list.item(i).text().strip() for i in range(self.ref_list.count()) if self.ref_list.item(i).text().strip()]
        samps_in_ui = [self.sample_list.item(i).text().strip() for i in range(self.sample_list.count()) if self.sample_list.item(i).text().strip()]
        
        def is_heco2_exception(name):
            clean = re.sub(r'[\s\-_]+', '', str(name).upper())
            return clean in ("HECO2", "CO2", "HEC02", "C02")

        mats = settings.get_setting("REFERENCE_MATERIALS", sub_key="Water") or []
        changed = False
        
        # 1. Prune items explicitly moved to samples
        new_mats = []
        for mat in mats:
            # If the setting material maps to any sample in the UI, we delete it from the table
            if any(self._is_mapped_to_existing(samp, [mat]) for samp in samps_in_ui):
                changed = True
            else:
                new_mats.append(mat)
        mats = new_mats
        
        # 2. Add unmapped reference items
        for ref in refs_in_ui:
            if is_heco2_exception(ref): continue # Ignore HeCO2
            if not self._is_mapped_to_existing(ref, mats):
                mats.append({"col_c": ref, "col_d": "", "col_e": "", "col_f": "", "col_g": "", "col_h": "", "color": "black", "bold": False})
                changed = True
                
        if changed:
            settings.set_setting("REFERENCE_MATERIALS", mats, sub_key="Water")
            self.references_updated.emit() # Triggers instant main_window refresh
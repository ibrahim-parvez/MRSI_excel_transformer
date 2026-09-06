import os
from copy import copy, deepcopy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell.rich_text import CellRichText, TextBlock
# Import CellIsRule, FormulaRule to ensure all imports are available for downstream use
from openpyxl.formatting.rule import CellIsRule, FormulaRule 
from utils.common_utils import embed_settings_popup, save_workbook_atomic
from utils.excel_engine import recalculate_workbook

def _is_formula_cell(cell):
    """Return True if the cell is a formula."""
    try:
        if getattr(cell, "data_type", None) == "f":
            return True
        val = cell.value
        return isinstance(val, str) and val.startswith("=")
    except Exception:
        return False

def _try_refresh_with_xlwings(path):
    """
    Force a full recalculation of `path` in Excel and save it.

    Delegates to the shared Excel engine so that, during a Combine run, this
    reuses that run's background Excel instance rather than starting and
    quitting its own (which on Windows would taskkill the Combine instance --
    see utils/excel_engine.py).
    """
    return recalculate_workbook(path, full=True)

def step7_report_carbonate(file_path):
    source_sheet = "Normalization_DNT"
    new_sheet_name = "Report_DNT"

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb_fmt = load_workbook(file_path, data_only=False)
    wb_val = load_workbook(file_path, data_only=True)

    if source_sheet not in wb_fmt.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found.")

    ws_fmt = wb_fmt[source_sheet]
    ws_val = wb_val[source_sheet]
    
    max_col = ws_fmt.max_column

    def _cell_rgb_upper(cell):
        try:
            fg = getattr(cell.fill, "fgColor", None)
            if fg is not None:
                rgb = getattr(fg, "rgb", None)
                if rgb:
                    return str(rgb).upper()
            sc = getattr(cell.fill, "start_color", None)
            if sc is not None:
                rgb2 = getattr(sc, "rgb", None)
                if rgb2:
                    return str(rgb2).upper()
        except Exception:
            pass
        return None

    def _is_gray808080(cell):
        rgb = _cell_rgb_upper(cell)
        return bool(rgb and rgb.endswith("808080"))

    # --- 1. Find the 2-Row Dark Gray Divider Band (color #808080) ---
    gray_band_start = None
    check_start_col = 19
    check_end_col = max_col # Dynamically bounded to the sheet's maximum populated column
    range_width = max(1, check_end_col - check_start_col + 1)
    threshold = max(1, range_width // 2)

    # Search method 1: Look for gray across columns 19 to end.
    for r in range(1, ws_fmt.max_row):
        count_r = sum(1 for c in range(check_start_col, check_end_col + 1)
                      if _is_gray808080(ws_fmt.cell(row=r, column=c)))
        count_r1 = sum(1 for c in range(check_start_col, check_end_col + 1)
                       if _is_gray808080(ws_fmt.cell(row=r + 1, column=c)))
        if count_r >= threshold and count_r1 >= threshold:
            gray_band_start = r
            break

    # Fallback search method 2: Look for gray across columns 12 to end of data.
    if gray_band_start is None:
        start_col_l = 12
        end_col = max_col
        width2 = max(1, end_col - start_col_l + 1)
        threshold2 = max(1, width2 // 2)
        for r in range(1, ws_fmt.max_row):
            count_r = sum(1 for c in range(start_col_l, end_col + 1)
                          if _is_gray808080(ws_fmt.cell(row=r, column=c)))
            count_r1 = sum(1 for c in range(start_col_l, end_col + 1)
                           if _is_gray808080(ws_fmt.cell(row=r + 1, column=c)))
            if count_r >= threshold2 and count_r1 >= threshold2:
                gray_band_start = r
                break

    # Fallback search method 3: Look for gray anywhere in two consecutive rows.
    if gray_band_start is None:
        for r in range(1, ws_fmt.max_row):
            any_r = any(_is_gray808080(ws_fmt.cell(row=r, column=c)) for c in range(1, max_col + 1))
            any_r1 = any(_is_gray808080(ws_fmt.cell(row=r + 1, column=c)) for c in range(1, max_col + 1)) if r < ws_fmt.max_row else False
            if any_r and any_r1:
                gray_band_start = r
                break

    if gray_band_start is None:
        raise ValueError("Could not find the 2-row dark gray band (color #808080) in 'Normalization_DNT' sheet.")

    start_row = max(1, gray_band_start - 3)
    
    # --- 2. Source columns for the dynamic layout ---
    # A=1, B=2, C=3 (Original data/identifier)
    # S=19 to end of sheet (Normalized Output data)
    source_cols = list(range(1, 4)) + list(range(19, max_col + 1)) 
    
    needs_refresh = False
    for r in range(start_row, min(start_row + 30, ws_fmt.max_row + 1)):
        for c in source_cols:
            src = ws_fmt.cell(row=r, column=c)
            valcell = ws_val.cell(row=r, column=c)
            if _is_formula_cell(src) and (valcell.value is None):
                needs_refresh = True
                break
        if needs_refresh:
            break

    if needs_refresh:
        refreshed = _try_refresh_with_xlwings(file_path)
        if refreshed:
            wb_fmt = load_workbook(file_path, data_only=False)
            wb_val = load_workbook(file_path, data_only=True)
            ws_fmt = wb_fmt[source_sheet]
            ws_val = wb_val[source_sheet]

    if new_sheet_name in wb_fmt.sheetnames:
        del wb_fmt[new_sheet_name]

    ws_new = wb_fmt.create_sheet(new_sheet_name, index=wb_fmt.index(ws_fmt))
    
    # The mapping converts source column index to the new, sequential column index.
    mapping = {src_col: idx for idx, src_col in enumerate(source_cols, start=1)}

    new_row = 1
    for r in range(start_row, ws_fmt.max_row + 1):
        for src_col in source_cols:
            new_col = mapping[src_col]
            src_cell_fmt = ws_fmt.cell(row=r, column=src_col)
            src_cell_val = ws_val.cell(row=r, column=src_col)
            dst = ws_new.cell(row=new_row, column=new_col)

            value = None
            if src_cell_val.value is not None:
                value = src_cell_val.value
            elif not _is_formula_cell(src_cell_fmt):
                value = src_cell_fmt.value

            # Copy rich text properly
            try:
                if hasattr(src_cell_fmt, "rich_text") and src_cell_fmt.rich_text:
                    rt = CellRichText()
                    for block in src_cell_fmt.rich_text:
                        if isinstance(block, TextBlock):
                            rt.append(deepcopy(block))
                    dst.rich_text = rt
                else:
                    dst.value = value
            except Exception:
                dst.value = value

            try:
                if getattr(src_cell_fmt, "comment", None) is not None:
                    dst.comment = deepcopy(src_cell_fmt.comment)
            except Exception:
                pass

            try:
                if src_cell_fmt.has_style:
                    dst.font = copy(src_cell_fmt.font)
                    dst.border = copy(src_cell_fmt.border)
                    dst.fill = copy(src_cell_fmt.fill)
                    dst.number_format = src_cell_fmt.number_format
                    dst.protection = copy(src_cell_fmt.protection)
                    dst.alignment = copy(src_cell_fmt.alignment)
            except Exception:
                pass

        try:
            rd = ws_fmt.row_dimensions.get(r)
            if rd is not None and getattr(rd, "height", None) is not None:
                ws_new.row_dimensions[new_row].height = rd.height
        except Exception:
            pass

        new_row += 1

    total_rows = new_row - 1

    # --- 3. RE-APPLY MERGED CELLS ---
    # openpyxl doesn't copy merged state natively, so we translate the coordinates
    for merged_range in ws_fmt.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = merged_range.bounds
        
        # Only process merges that fall within the rows we actually copied
        if min_row >= start_row:
            # Ensure the merged columns are part of the columns we mapped over
            if min_col in mapping and max_col_m in mapping:
                new_min_col = mapping[min_col]
                new_max_col = mapping[max_col_m]
                
                # Shift the rows up based on where we started copying
                new_min_row = min_row - start_row + 1
                new_max_row = max_row_m - start_row + 1
                
                try:
                    ws_new.merge_cells(start_row=new_min_row, start_column=new_min_col,
                                       end_row=new_max_row, end_column=new_max_col)
                except Exception:
                    pass

    for src_col, new_col in mapping.items():
        try:
            src_letter = get_column_letter(src_col)
            new_letter = get_column_letter(new_col)
            cd = ws_fmt.column_dimensions.get(src_letter)
            if cd is not None and getattr(cd, "width", None) is not None:
                ws_new.column_dimensions[new_letter].width = cd.width
        except Exception:
            pass

    # Set Summary to open at A1 and be active
    for s in wb_fmt.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws_new.sheet_view.tabSelected = True
    wb_fmt.active = wb_fmt.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(ws_new, "A1")

    save_workbook_atomic(wb_fmt, file_path)
    print(f"✅ Step 7: Report completed on {file_path}")
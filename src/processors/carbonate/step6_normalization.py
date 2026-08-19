import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy
import re
import unicodedata
import statistics
import utils.settings as settings
from utils.common_utils import embed_settings_popup
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# --- Helper Functions ---
def _normalize_text(text):
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r"[^A-Za-z0-9]+", "", text)
    return text.lower().strip()

def create_rich_text(parts):
    rt = CellRichText()
    for font, text in parts:
        rt.append(TextBlock(font, text))
    return rt

def get_subscript_label(main_text, sub_text, base_font=None):
    """Creates a CellRichText object with subscripted text."""
    if base_font is None:
        base_font = InlineFont(rFont="Calibri", sz=11, color="000000")
    sub_font = copy(base_font)
    sub_font.vertAlign = "subscript"
    return create_rich_text([
        (base_font, main_text),
        (sub_font, sub_text)
    ])

def get_co2_temp_label():
    """Retrieves dynamic temperature text from settings for CO2 normalization."""
    # We now fetch the exact string saved by the UI
    mode = settings.get_setting("CO2_TEMP_MODE")
    
    if mode == "72 °C":
        return "Using 72 °C"
    elif mode == "25 °C":
        return "Using 25 °C"
    elif mode == "Custom":
        return "Using Custom Temp"
        
    # Absolute fallback just in case settings fail to load completely
    return "Using 25 °C"

def extract_sample_base(identifier):
    if not identifier or not isinstance(identifier, str):
        return ""
    identifier = identifier.strip()
    base = re.sub(r"\s*r\d+(\.\d+)?$", "", identifier, flags=re.IGNORECASE)
    return base.strip()

def extract_run_number(identifier):
    if not identifier or not isinstance(identifier, str):
        return (9999, 0)
    m = re.search(r"r(\d+)(?:\.(\d+))?", identifier, flags=re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (9999, 0)

def _make_fill(hex_color):
    c = hex_color.replace("#", "").upper()
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def get_excel_range(col_let, rows):
    """Converts a list of row numbers into a valid Excel contiguous or multi-area range string."""
    if not rows:
        return ""
    rows = sorted(list(set(rows)))
    blocks = []
    start = rows[0]
    prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
        else:
            blocks.append((start, prev))
            start = r
            prev = r
    blocks.append((start, prev))
    
    range_strs = []
    for s, e in blocks:
        if s == e:
            range_strs.append(f"{col_let}{s}")
        else:
            range_strs.append(f"{col_let}{s}:{col_let}{e}")
            
    if len(range_strs) == 1:
        return range_strs[0]
    else:
        return f"_xlfn.VSTACK({','.join(range_strs)})"

def _get_valid_co2_rows(rows, col_identifier1):
    valid_indices = []
    seen = {}
    for i, r in enumerate(rows):
        ident_raw = str(r[col_identifier1 - 1] or "").strip()
        ident_lower = ident_raw.lower()
        if ident_lower.startswith("heco2"):
            ident_clean = "co2" + ident_raw[len("heco2"):]
        elif ident_lower.startswith("co2"):
            ident_clean = "co2" + ident_raw[len("co2"):]
        else:
            ident_clean = ident_raw
        major, minor = extract_run_number(ident_clean)
        if major is None or minor is None:
            continue
        if major == 1:
            continue
        if major not in seen:
            seen[major] = (minor, i)
        else:
            prev_minor, prev_i = seen[major]
            if minor < prev_minor:
                seen[major] = (minor, i)
    return sorted(idx for _, idx in seen.values())

def get_summary_num_format(base_name):
    return '0.000'

def draw_blue_box_structure(ws, offset_col=0, setting_key="Carbonate", box_fill=None):
    if not box_fill:
        box_fill = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
        
    if offset_col > 0:
        for r in range(1, 25):
            for c in range(3 + offset_col, 16 + offset_col):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.border = Border()
                cell.fill = PatternFill()

    materials = settings.get_setting("REFERENCE_MATERIALS", sub_key=setting_key)
    if not materials: materials = []
    
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key=setting_key)
    if not slope_groups: slope_groups = []
    
    num_materials = len(materials)
    
    data_start_row = 4
    data_end_row = 4 + num_materials - 1 if num_materials > 0 else 4

    thick = Side(border_style="thick", color="000000")
    medium = Side(border_style="medium", color="000000")
    black_bold = Font(color="000000", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    def get_style_font(color_name, is_bold):
        colors = {"black": "000000", "green": "008000", "red": "FF0000", 
                  "darkblue": "000080", "lightblue": "3399FF", "orange": "FF9900"}
        hex_code = colors.get(str(color_name).lower(), "000000")
        return Font(color=hex_code, bold=is_bold)

    today_str = datetime.today().strftime("%Y-%m-%d")
    
    if offset_col == 0:
        ws.cell(row=1, column=1, value=today_str).alignment = Alignment(horizontal="left", vertical="center")
        header_title = f"Normalization ({setting_key})"
    else:
        temp_note = get_co2_temp_label()
        header_title = f"Normalization ({setting_key}) - {temp_note}"
        
    ws.cell(row=1, column=3 + offset_col, value=header_title).font = black_bold
    ws.cell(row=1, column=3 + offset_col).alignment = center
    
    col_start = 3 + offset_col; col_end = 8 + offset_col
    ws.merge_cells(start_row=2, start_column=3 + offset_col, end_row=3, end_column=3 + offset_col)
    ws.cell(row=2, column=3 + offset_col, value="Reference Materials").font = black_bold
    ws.cell(row=2, column=3 + offset_col).alignment = center
    ws.merge_cells(start_row=2, start_column=6 + offset_col, end_row=2, end_column=7 + offset_col)
    ws.cell(row=2, column=6 + offset_col, value="Published (vs. VPDB)").font = black_bold
    ws.cell(row=2, column=6 + offset_col).alignment = center

    # Setup Subscripts for Carbonate vs CO2
    if setting_key == "Carbonate":
        ws.cell(row=3, column=6 + offset_col).value = get_subscript_label("δ¹³C", "Carbonate")
        ws.cell(row=3, column=7 + offset_col, value="δ¹⁸O")
    else:
        ws.cell(row=3, column=6 + offset_col, value="δ¹³C")
        ws.cell(row=3, column=7 + offset_col).value = get_subscript_label("δ¹⁸O", "CO2")

    ws.cell(row=3, column=6 + offset_col).alignment = center
    ws.cell(row=3, column=7 + offset_col).alignment = center

    for r in range(2, 4):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = box_fill
            cell.border = Border(top=medium if r==2 else None, bottom=medium if r==3 else None, 
                                 left=medium if c==col_start else None, right=medium if c==col_end else None)

    active_slope_materials = []
    for group in slope_groups:
        for item in group:
            active_slope_materials.append(str(item).strip().lower())

    for idx, mat in enumerate(materials):
        r = data_start_row + idx
        mat_name = str(mat.get("col_c", "")).strip().lower()
        mat_clean = re.sub(r'[\s\-_]+', '', mat_name.upper())
        mat_no_std = mat_clean.replace("STD", "")
        
        is_used_in_slope = False
        for active_mat in active_slope_materials:
            active_clean = re.sub(r'[\s\-_]+', '', active_mat.upper())
            active_no_std = active_clean.replace("STD", "")
            if active_clean in mat_clean or (len(active_no_std) >= 4 and active_no_std in mat_no_std):
                is_used_in_slope = True
                break
                
        c_val = mat.get("col_c")
        d_val = mat.get("col_d")
        
        e_raw = mat.get("col_e")
        f_raw = mat.get("col_f")
        g_raw = mat.get("col_g")
        h_raw = mat.get("col_h")
        
        d13c_val = f_raw if (f_raw is not None and str(f_raw).strip() != "") else e_raw
        d18o_val = g_raw if (g_raw is not None and str(g_raw).strip() != "") else h_raw
        
        if is_used_in_slope:
            e_val = None
            f_val = d13c_val
            g_val = d18o_val
            h_val = None
        else:
            e_val = d13c_val
            f_val = None
            g_val = None
            h_val = d18o_val

        vals = [c_val, d_val, e_val, f_val, g_val, h_val]
        font_style = get_style_font(mat.get("color", "black"), mat.get("bold", False))
        
        for i, val in enumerate(vals):
            c = col_start + i
            if i in [2, 3, 4, 5]: 
                try: val = float(val)
                except: pass
                
            target_cell = ws.cell(row=r, column=c, value=val)
            target_cell.font = font_style
            target_cell.alignment = center
            
            if i in [2, 3, 4, 5] and isinstance(val, (int, float)):
                target_cell.number_format = '0.00'
            
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = box_fill
            cell.border = Border(bottom=medium if r==data_end_row else None, 
                                 left=medium if c==col_start else None, right=medium if c==col_end else None)

    ws.merge_cells(start_row=2, start_column=10 + offset_col, end_row=2, end_column=14 + offset_col)
    ws.cell(row=2, column=10 + offset_col, value="Measured (vs. Working Standard)").font = black_bold
    ws.cell(row=2, column=10 + offset_col).alignment = center

    if setting_key == "Carbonate":
        ws.cell(row=3, column=11 + offset_col).value = get_subscript_label("δ¹³C", "Carbonate")
        ws.cell(row=3, column=14 + offset_col, value="δ¹⁸O")
    else:
        ws.cell(row=3, column=11 + offset_col, value="δ¹³C")
        ws.cell(row=3, column=14 + offset_col).value = get_subscript_label("δ¹⁸O", "CO2")

    ws.cell(row=3, column=11 + offset_col).alignment = center
    ws.cell(row=3, column=14 + offset_col).alignment = center
    
    for r in range(2, data_end_row + 1):
        for c in range(10 + offset_col, 15 + offset_col):
            cell = ws.cell(row=r, column=c)
            cell.fill = box_fill
            cell.alignment = center
            top = thick if r == 2 else None
            bottom = thick if r == data_end_row else None
            left = thick if c == 10 + offset_col else None
            right = thick if c == 14 + offset_col else None
            if r == 3: bottom = thick
            if r == 4: top = None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    red_if = InlineFont(color='00FF0000', b=True)
    blue_if = InlineFont(color='000000FF', b=True)
    green_if = InlineFont(color='008000', b=True)
    black_if = InlineFont(color='000000', b=True)

    def get_group_rich_text(group_list):
        parts = []
        for item in group_list:
            clean = str(item).strip().upper()
            font = black_if
            if "18" in clean: font = red_if
            elif "19" in clean: font = blue_if
            elif "603" in clean: font = green_if
            parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
        return create_rich_text(parts)

    current_row = data_end_row + 2 
    slope_info = []

    for group_list in slope_groups:
        ws.cell(row=current_row, column=10 + offset_col, value="Slope").font = black_bold
        ws.cell(row=current_row + 1, column=10 + offset_col, value="Intercept").font = black_bold
        
        ws.cell(row=current_row, column=9 + offset_col).value = get_group_rich_text(group_list)
        ws.cell(row=current_row, column=9 + offset_col).alignment = center
        
        for r_off in [0, 1]:
            ws.cell(row=current_row+r_off, column=11 + offset_col).alignment = center
            ws.cell(row=current_row+r_off, column=14 + offset_col).alignment = center

        slope_info.append({"slope_row": current_row, "intercept_row": current_row + 1})
        current_row += 3 

    mat_row_map = {}
    for idx, mat in enumerate(materials):
        if mat.get("col_c"): mat_row_map[str(mat.get("col_c")).strip().lower()] = data_start_row + idx

    return current_row - 1, current_row + 3, slope_info, mat_row_map

def populate_blue_box_math(ws, slope_info, mat_row_map, offset_col=0, setting_key="Carbonate"):
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key=setting_key)
    if not slope_groups: slope_groups = []
    
    calc_mode = settings.get_setting("CALC_MODE_STEP7")
    use_outlier_excluded = (calc_mode == "Outliers Excluded")

    materials = settings.get_setting("REFERENCE_MATERIALS", sub_key=setting_key)
    if not materials: materials = []

    def get_style_font(color_name, is_bold):
        colors = {"black": "000000", "green": "008000", "red": "FF0000", 
                  "darkblue": "000080", "lightblue": "3399FF", "orange": "FF9900"}
        hex_code = colors.get(str(color_name).lower(), "000000")
        return Font(color=hex_code, bold=is_bold)
    
    identifier_col = 3
    c_avg_col_ref = 11 
    o_avg_col_ref = 14 
    
    found_map = {} 
    
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=c_avg_col_ref).value
        if val and str(val).strip().lower() == "average":
            label_val = str(ws.cell(row=r, column=10).value or "").strip()
            is_excluded_block = ("outlier" in label_val.lower())
            
            if use_outlier_excluded and not is_excluded_block: continue
            if not use_outlier_excluded and is_excluded_block: continue
                
            avg_row = r + 1 
            id_row = r - 1
            ident = ""
            for t in range(id_row, max(1, id_row - 25), -1):
                cellv = ws.cell(row=t, column=identifier_col).value
                if cellv:
                    ident = str(cellv).strip().lower()
                    break
            
            ident_clean = re.sub(r'[\s\-_]+', '', ident.upper())
            ident_no_std = ident_clean.replace("STD", "")
            
            for mat_name, target_row in mat_row_map.items():
                orig_mat = next((m for m in materials if str(m.get("col_c","")).strip().lower() == mat_name), None)
                if not orig_mat: continue
                
                std_clean = re.sub(r'[\s\-_]+', '', str(orig_mat.get("col_c", "")).upper())
                std_no_std = std_clean.replace("STD", "")
                
                is_match = False
                if std_clean in ident_clean:
                    is_match = True
                elif len(std_no_std) >= 4 and std_no_std in ident_no_std:
                    is_match = True
                    
                if is_match:
                    color = orig_mat.get("color", "black")
                    bold = orig_mat.get("bold", False)
                    font_style = get_style_font(color, bold)

                    c_cell = ws.cell(row=target_row, column=11 + offset_col, value=f'=IFERROR({get_column_letter(c_avg_col_ref)}{avg_row},"")')
                    c_cell.font = font_style
                    
                    o_cell = ws.cell(row=target_row, column=14 + offset_col, value=f'=IFERROR({get_column_letter(o_avg_col_ref)}{avg_row},"")')
                    o_cell.font = font_style
                    
                    found_map[mat_name] = target_row
                    break

    for idx, group_list in enumerate(slope_groups):
        if idx >= len(slope_info): break 
        
        current_slope_row = slope_info[idx]["slope_row"]
        
        rows = []
        for n in group_list:
            key = str(n).strip().lower()
            if key in mat_row_map and key in found_map:
                rows.append(mat_row_map[key])
        
        if len(rows) >= 2:
            range_y_pub = get_excel_range(get_column_letter(6 + offset_col), rows)
            range_x_meas = get_excel_range(get_column_letter(11 + offset_col), rows)
            
            range_o_pub = get_excel_range(get_column_letter(7 + offset_col), rows)
            range_o_meas = get_excel_range(get_column_letter(14 + offset_col), rows)
            
            ws.cell(row=current_slope_row, column=11 + offset_col, value=f'=SLOPE({range_y_pub},{range_x_meas})')
            ws.cell(row=current_slope_row+1, column=11 + offset_col, value=f'=INTERCEPT({range_y_pub},{range_x_meas})')
            ws.cell(row=current_slope_row, column=14 + offset_col, value=f'=SLOPE({range_o_pub},{range_o_meas})')
            ws.cell(row=current_slope_row+1, column=14 + offset_col, value=f'=INTERCEPT({range_o_pub},{range_o_meas})')

def draw_yield_table(ws, start_col, box_fill, num_yield_groups):
    max_yield_row = 19 + (num_yield_groups * 3)
    for r in range(1, max_yield_row + 1):
        for c in range(start_col, start_col + 13):
            ws.cell(row=r, column=c).fill = box_fill

    black_bold = Font(color="000000", bold=True)
    green_bold = Font(color="00B050", bold=True)
    thick = Side(border_style="medium", color="000000") 
    
    ws.cell(row=1, column=start_col, value="By Sang-Tae Kim (2026-0726)").font = black_bold
    ws.cell(row=2, column=start_col + 4, value="Yield Calculation (Sample weight vs. Sum area all)").font = black_bold
    
    ah_col = start_col + 1
    ai_col = start_col + 2
    
    col_AI = get_column_letter(start_col + 2)
    col_AL = get_column_letter(start_col + 5)
    col_AN = get_column_letter(start_col + 7)
    col_AP = get_column_letter(start_col + 9)
    col_AQ = get_column_letter(start_col + 10)
    
    ws.cell(row=2, column=ah_col, value="Atomic weight").font = black_bold
    ws.cell(row=2, column=ah_col).alignment = Alignment(horizontal="center")
    
    atomic_data = [
        ("Mn", 54.93805),
        ("Fe", 55.845),
        ("Li", 6.941),
        ("Ca", 40.078),
        ("Mg", 24.3050),
        ("C", 12.0107),
        ("O", 15.9994)
    ]
    
    for i, (elem, weight) in enumerate(atomic_data, start=3):
        ws.cell(row=i, column=ah_col, value=elem)
        cell_wt = ws.cell(row=i, column=ai_col, value=weight)
        cell_wt.number_format = '0.00E+00'
        
    ws.cell(row=10, column=ah_col, value="CO3")
    ws.cell(row=10, column=ai_col, value=f"={col_AI}8+({col_AI}9*3)").number_format = '0.00E+00'
    ws.cell(row=11, column=ah_col, value="CO2")
    ws.cell(row=11, column=ai_col, value=f"={col_AI}8+({col_AI}9*2)").number_format = '0.00E+00'
    
    ws.cell(row=12, column=ai_col, value="Mol. Weight").font = green_bold
    ws.cell(row=12, column=start_col + 3, value="CO2 weight %").font = green_bold 
    ws.cell(row=12, column=start_col + 5, value="Yields (umol/mg)").font = green_bold 
    ws.cell(row=12, column=start_col + 7, value="Yields (mmol/mg)").font = green_bold 
    
    ws.cell(row=11, column=start_col + 9, value="Target").font = black_bold
    ws.cell(row=11, column=start_col + 9).alignment = Alignment(horizontal="center")
    ws.cell(row=12, column=start_col + 9, value="mg").font = black_bold
    ws.cell(row=12, column=start_col + 9).alignment = Alignment(horizontal="center")
    ws.cell(row=12, column=start_col + 10, value="mmol").font = black_bold
    ws.cell(row=12, column=start_col + 10).alignment = Alignment(horizontal="center")
    ws.cell(row=12, column=start_col + 11, value="umol").font = black_bold
    ws.cell(row=12, column=start_col + 11).alignment = Alignment(horizontal="center")
    
    # --- Fetch Dynamic Settings for Yield ---
    yield_compounds = settings.get_setting("YIELD_COMPOUNDS") or {"ref": "CaCO3", "samp": "MnCO3"}
    ref_comp = yield_compounds.get("ref", "CaCO3")
    samp_comp = yield_compounds.get("samp", "MnCO3")
    
    base_colors = {
        "CaCO3": "00B050", 
        "MgCO3": "00B050",
        "CaMg(CO3)2": "006633",
        "Li2CO3": "33CC33",
        "MnCO3": "00B050", 
        "FeCO3": "339966"
    }

    def get_fonts(comp_name):
        if comp_name == ref_comp:
            return Font(color="000000", bold=True), Font(color="000000")
        elif comp_name == samp_comp:
            return Font(color="FF66B2", bold=True), Font(color="FF66B2")
        else:
            return Font(color=base_colors.get(comp_name, "00B050"), bold=True), Font(color="00B050")

    f_name_ca, f_val_ca = get_fonts("CaCO3")
    f_name_mg, f_val_mg = get_fonts("MgCO3")
    f_name_camg, f_val_camg = get_fonts("CaMg(CO3)2")
    f_name_li, f_val_li = get_fonts("Li2CO3")
    f_name_mn, f_val_mn = get_fonts("MnCO3")
    f_name_fe, f_val_fe = get_fonts("FeCO3")
    
    compounds = [
        (13, "CaCO3", f_name_ca, f_val_ca, f"={col_AI}6+{col_AI}10", f"=(${col_AI}$11/{col_AI}13)*100", f"=1000/{col_AI}13", f"={col_AL}13/1000", 0.150),
        (14, "MgCO3", f_name_mg, f_val_mg, f"={col_AI}7+{col_AI}10", f"=(${col_AI}$11/{col_AI}14)*100", f"=1000/{col_AI}14", f"={col_AL}14/1000", 0.126),
        (15, "CaMg(CO3)2", f_name_camg, f_val_camg, f"={col_AI}6+{col_AI}7+(2*{col_AI}10)", f"=(${col_AI}$11)*2/{col_AI}15*100", f"=(1000/{col_AI}15)*2", f"={col_AL}15/1000", 0.138),
        (16, "Li2CO3", f_name_li, f_val_li, f"=({col_AI}5*2)+{col_AI}10", f"=(${col_AI}$11/{col_AI}16)*100", f"=1000/{col_AI}16", f"={col_AL}16/1000", None),
        (17, "MnCO3", f_name_mn, f_val_mn, f"={col_AI}3+{col_AI}10", f"=(${col_AI}$11/{col_AI}17)*100", f"=1000/{col_AI}17", f"={col_AL}17/1000", 0.172),
        (18, "FeCO3", f_name_fe, f_val_fe, f"={col_AI}4+{col_AI}10", f"=(${col_AI}$11/{col_AI}18)*100", f"=1000/{col_AI}18", f"={col_AL}18/1000", 0.174),
    ]
    
    sci_format = '0.00E+00'
    for r, name, name_font, val_font, mol_wt, co2_wt, y_umol, y_mmol, target_mg in compounds:
        ws.cell(row=r, column=ah_col, value=name).font = name_font
        ws.cell(row=r, column=ai_col, value=mol_wt).font = val_font; ws.cell(row=r, column=ai_col).number_format = sci_format
        ws.cell(row=r, column=start_col + 3, value=co2_wt).font = val_font; ws.cell(row=r, column=start_col + 3).number_format = sci_format
        ws.cell(row=r, column=start_col + 5, value=y_umol).font = val_font; ws.cell(row=r, column=start_col + 5).number_format = sci_format
        ws.cell(row=r, column=start_col + 7, value=y_mmol).font = val_font; ws.cell(row=r, column=start_col + 7).number_format = sci_format
        
        if target_mg is not None:
            ws.cell(row=r, column=start_col + 9, value=target_mg).number_format = sci_format
            ws.cell(row=r, column=start_col + 10, value=f"={col_AP}{r}*{col_AN}{r}").number_format = sci_format
            ws.cell(row=r, column=start_col + 11, value=f"={col_AQ}{r}*1000").number_format = sci_format
            
    for r in range(2, 12):
        for c in range(ah_col, ai_col + 1):
            border = Border(
                top=thick if r == 2 else None,
                bottom=thick if r == 11 or r == 2 else None,
                left=thick if c == ah_col else None,
                right=thick if c == ai_col else None
            )
            ws.cell(row=r, column=c).border = border
            
    for r in range(12, 19):
        for c in range(ah_col, start_col + 8):
            border = Border(
                top=thick if r == 12 else None,
                bottom=thick if r == 18 or r == 12 else None,
                left=thick if c == ah_col else None,
                right=thick if c == start_col + 7 else None
            )
            ws.cell(row=r, column=c).border = border

def draw_upper_boxes(ws, h_row, box_fill, black_bold, green_bold, output_offset=0, setting_key="Carbonate"):
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key=setting_key) or []
    num_groups = len(slope_groups)
    if num_groups == 0:
        return
    
    c_start = 19 + output_offset
    o_calc_start = c_start + num_groups + 1 
    o_arag_start = o_calc_start + num_groups + 1
    
    box1_start = c_start
    box1_end = o_arag_start + num_groups - 1
    
    box2_start = box1_end + 2
    vsmow_calc_start = box2_start
    vsmow_arag_start = vsmow_calc_start + num_groups + 1 
    box2_end = vsmow_arag_start + num_groups - 1

    thick = Side(style="thick")
    
    red_font = InlineFont(color='00FF0000', b=True)
    blue_font = InlineFont(color='000000FF', b=True)
    green_font = InlineFont(color='008000', b=True)
    black_font = InlineFont(color='000000', b=True)
    
    def get_rich_text_for_group(group_list):
        parts = []
        for item in group_list:
            clean = str(item).strip().upper()
            font = black_font
            if "18" in clean: font = red_font
            elif "19" in clean: font = blue_font
            elif "603" in clean: font = green_font
            parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
        return create_rich_text(parts)
        
    for r in range(h_row - 3, h_row + 1):
        for c in range(box1_start, box1_end + 1):
            cell = ws.cell(r, c)
            cell.fill = box_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            top = thick if r == h_row - 3 else None
            bottom = thick if r == h_row else None 
            left = thick if c == box1_start else None
            right = thick if c == box1_end else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.merge_cells(start_row=h_row - 3, start_column=box1_start, end_row=h_row - 3, end_column=box1_end)
    ws.cell(h_row - 3, box1_start, f"Normalized VPDB ({setting_key})").font = black_bold

    if num_groups > 1:
        ws.merge_cells(start_row=h_row - 2, start_column=o_calc_start, end_row=h_row - 2, end_column=o_calc_start + num_groups - 1)
        ws.merge_cells(start_row=h_row - 2, start_column=o_arag_start, end_row=h_row - 2, end_column=box1_end)
        
    ws.cell(h_row - 2, o_calc_start, "Calcite").font = black_bold
    ws.cell(h_row - 2, o_arag_start, "Aragonite").font = green_bold

    for i, grp in enumerate(slope_groups):
        rt = get_rich_text_for_group(grp)
        if setting_key == "Carbonate":
            ws.cell(h_row - 1, c_start + i).value = get_subscript_label("δ¹³C", "Carbonate")
            ws.cell(h_row - 1, o_calc_start + i, "δ¹⁸O")
            ws.cell(h_row - 1, o_arag_start + i, "δ¹⁸O")
        else:
            ws.cell(h_row - 1, c_start + i, "δ¹³C")
            ws.cell(h_row - 1, o_calc_start + i).value = get_subscript_label("δ¹⁸O", "CO2")
            ws.cell(h_row - 1, o_arag_start + i).value = get_subscript_label("δ¹⁸O", "CO2")

        ws.cell(h_row, c_start + i).value = rt
        ws.cell(h_row, o_calc_start + i).value = rt
        ws.cell(h_row, o_arag_start + i).value = rt

    for r in range(h_row - 3, h_row + 1):
        for c in range(box2_start, box2_end + 1):
            cell = ws.cell(r, c)
            cell.fill = box_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            top = thick if r == h_row - 3 else None
            bottom = thick if r == h_row else None
            left = thick if c == box2_start else None
            right = thick if c == box2_end else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.merge_cells(start_row=h_row - 3, start_column=box2_start, end_row=h_row - 3, end_column=box2_end)
    ws.cell(h_row - 3, box2_start, "VSMOW").font = black_bold

    if num_groups > 1:
        ws.merge_cells(start_row=h_row - 2, start_column=vsmow_calc_start, end_row=h_row - 2, end_column=vsmow_calc_start + num_groups - 1)
        ws.merge_cells(start_row=h_row - 2, start_column=vsmow_arag_start, end_row=h_row - 2, end_column=box2_end)
        
    ws.cell(h_row - 2, vsmow_calc_start, "Calcite").font = black_bold
    ws.cell(h_row - 2, vsmow_arag_start, "Aragonite").font = green_bold
    
    for i, grp in enumerate(slope_groups):
        rt = get_rich_text_for_group(grp)
        if setting_key == "Carbonate":
            ws.cell(h_row - 1, vsmow_calc_start + i, "δ¹⁸O")
            ws.cell(h_row - 1, vsmow_arag_start + i, "δ¹⁸O")
        else:
            ws.cell(h_row - 1, vsmow_calc_start + i).value = get_subscript_label("δ¹⁸O", "CO2")
            ws.cell(h_row - 1, vsmow_arag_start + i).value = get_subscript_label("δ¹⁸O", "CO2")

        ws.cell(h_row, vsmow_calc_start + i).value = rt
        ws.cell(h_row, vsmow_arag_start + i).value = rt

def draw_lower_boxes(ws, divider_top_row, box_fill, black_bold, green_bold, output_offset=0, setting_key="Carbonate"):
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key=setting_key) or []
    num_groups = len(slope_groups)
    if num_groups == 0:
        return
    
    c_start = 19 + output_offset
    o_calc_start = c_start + num_groups + 1 
    o_arag_start = o_calc_start + num_groups + 1
    
    box1_start = c_start
    box1_end = o_arag_start + num_groups - 1
    
    box2_start = box1_end + 2
    vsmow_calc_start = box2_start
    vsmow_arag_start = vsmow_calc_start + num_groups + 1 
    box2_end = vsmow_arag_start + num_groups - 1

    thick = Side(style="thick")
    
    for r in range(divider_top_row - 3, divider_top_row + 2):
        for c in range(box1_start, box1_end + 1):
            cell = ws.cell(r, c)
            cell.fill = box_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            top = thick if r == divider_top_row - 3 else None
            bottom = thick if r == divider_top_row + 1 else None 
            left = thick if c == box1_start else None
            right = thick if c == box1_end else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.cell(divider_top_row - 3, box1_start, f"Normalized ({setting_key})").font = black_bold
    ws.cell(divider_top_row - 2, box1_start, "VPDB").font = black_bold
    
    if num_groups > 1:
        ws.merge_cells(start_row=divider_top_row - 1, start_column=o_calc_start, end_row=divider_top_row - 1, end_column=o_calc_start + num_groups - 1)
        ws.merge_cells(start_row=divider_top_row - 1, start_column=o_arag_start, end_row=divider_top_row - 1, end_column=box1_end)
        
    ws.cell(divider_top_row - 1, o_calc_start, "Calcite").font = black_bold
    ws.cell(divider_top_row - 1, o_arag_start, "Aragonite").font = green_bold
    
    red_font = InlineFont(color='00FF0000', b=True)
    blue_font = InlineFont(color='000000FF', b=True)
    green_font = InlineFont(color='008000', b=True)
    black_font = InlineFont(color='000000', b=True)
    
    def get_rich_text_for_group(group_list):
        parts = []
        for item in group_list:
            clean = str(item).strip().upper()
            font = black_font
            if "18" in clean: font = red_font
            elif "19" in clean: font = blue_font
            elif "603" in clean: font = green_font
            parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
        return create_rich_text(parts)
        
    for i, grp in enumerate(slope_groups):
        rt = get_rich_text_for_group(grp)
        if setting_key == "Carbonate":
            ws.cell(divider_top_row, c_start + i).value = get_subscript_label("δ¹³C", "Carbonate")
            ws.cell(divider_top_row, o_calc_start + i, "δ¹⁸O")
            ws.cell(divider_top_row, o_arag_start + i, "δ¹⁸O")
        else:
            ws.cell(divider_top_row, c_start + i, "δ¹³C")
            ws.cell(divider_top_row, o_calc_start + i).value = get_subscript_label("δ¹⁸O", "CO2")
            ws.cell(divider_top_row, o_arag_start + i).value = get_subscript_label("δ¹⁸O", "CO2")

        ws.cell(divider_top_row + 1, c_start + i).value = rt
        ws.cell(divider_top_row + 1, o_calc_start + i).value = rt
        ws.cell(divider_top_row + 1, o_arag_start + i).value = rt

    for r in range(divider_top_row - 2, divider_top_row + 2):
        for c in range(box2_start, box2_end + 1):
            cell = ws.cell(r, c)
            cell.fill = box_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            top = thick if r == divider_top_row - 2 else None
            bottom = thick if r == divider_top_row + 1 else None
            left = thick if c == box2_start else None
            right = thick if c == box2_end else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.cell(divider_top_row - 2, box2_start, "VSMOW").font = black_bold
    
    if num_groups > 1:
        ws.merge_cells(start_row=divider_top_row - 1, start_column=vsmow_calc_start, end_row=divider_top_row - 1, end_column=vsmow_calc_start + num_groups - 1)
        ws.merge_cells(start_row=divider_top_row - 1, start_column=vsmow_arag_start, end_row=divider_top_row - 1, end_column=box2_end)
        
    ws.cell(divider_top_row - 1, vsmow_calc_start, "Calcite").font = black_bold
    ws.cell(divider_top_row - 1, vsmow_arag_start, "Aragonite").font = green_bold
    
    for i, grp in enumerate(slope_groups):
        rt = get_rich_text_for_group(grp)
        if setting_key == "Carbonate":
            ws.cell(divider_top_row, vsmow_calc_start + i, "δ¹⁸O")
            ws.cell(divider_top_row, vsmow_arag_start + i, "δ¹⁸O")
        else:
            ws.cell(divider_top_row, vsmow_calc_start + i).value = get_subscript_label("δ¹⁸O", "CO2")
            ws.cell(divider_top_row, vsmow_arag_start + i).value = get_subscript_label("δ¹⁸O", "CO2")

        ws.cell(divider_top_row + 1, vsmow_calc_start + i).value = rt
        ws.cell(divider_top_row + 1, vsmow_arag_start + i).value = rt

def _detect_decimal_places_from_format(fmt: str):
    if not fmt or not isinstance(fmt, str):
        return None
    first_section = fmt.split(';', 1)[0]
    if '.' not in first_section:
        return None
    after_dot = first_section.split('.', 1)[1]
    count = 0
    for ch in after_dot:
        if ch in ('0', '#'):
            count += 1
        else:
            break
    return count if count > 0 else None


# --- Main Function ---
def step6_normalization_carbonate(file_path):
    
    carb_slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key="Carbonate") or []
    num_slope_groups = len(carb_slope_groups)
    
    EXTRA_HORIZ_SHIFT = max(0, num_slope_groups - 2) * 5
    
    yield_slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key="Yield") or []
    num_yield_groups = len(yield_slope_groups)
    
    EXTRA_VERT_SHIFT = max(0, num_yield_groups - 2) * 3
    
    # --- Fetch UI settings toggles ---
    calc_yield = settings.get_setting("CALC_YIELD")
    if calc_yield is None: calc_yield = True  # Default fallback
    
    calc_co2 = settings.get_setting("CALC_CO2")
    if calc_co2 is None: calc_co2 = True      # Default fallback

    # --- GLOBAL SHIFT VARIABLES ---
    YIELD_START_COL = 36 + EXTRA_HORIZ_SHIFT
    
    if calc_yield:
        CO2_TOP_SHIFT = 57 + EXTRA_HORIZ_SHIFT   
        CO2_OUT_SHIFT = 40 + EXTRA_HORIZ_SHIFT   
    else:
        # Shifts CO2 left to seamlessly occupy the empty Yield space
        CO2_TOP_SHIFT = 34 + EXTRA_HORIZ_SHIFT   
        CO2_OUT_SHIFT = 17 + EXTRA_HORIZ_SHIFT   
        
    HEADER_ROW_SHIFT = 12 
    # ------------------------------

    # --- Fetch Yield Compounds and Map to Excel Rows ---
    yield_compounds = settings.get_setting("YIELD_COMPOUNDS") or {"ref": "CaCO3", "samp": "MnCO3"}
    compound_row_map = {
        "CaCO3": 13, 
        "MgCO3": 14, 
        "CaMg(CO3)2": 15, 
        "Li2CO3": 16, 
        "MnCO3": 17, 
        "FeCO3": 18
    }
    ref_compound_row = compound_row_map.get(yield_compounds.get("ref", "CaCO3"), 13)
    samp_compound_row = compound_row_map.get(yield_compounds.get("samp", "MnCO3"), 17)

    stdev_is_enabled = settings.get_setting("STDEV_THRESHOLD_ENABLED")
    
    if stdev_is_enabled:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    else:
        stdev_threshold = None
    outlier_sigma = settings.get_setting("OUTLIER_SIGMA") or 2
    exclusion_mode = settings.get_setting("OUTLIER_EXCLUSION_MODE") or "Individual"

    strike_font = Font(strike=True, color="FF0000")

    EXCLUDED_COLS = {8, 9, 11, 12, 13, 26, 27, 14, 15}
    MAX_SOURCE_COL = 24
    
    col_map = {}
    dest_col = 1
    for src_col in range(1, MAX_SOURCE_COL + 1):
        if src_col not in EXCLUDED_COLS:
            col_map[src_col] = dest_col
            dest_col += 1
    
    col_identifier1 = col_map.get(3, 0)
    if col_identifier1 != 3:
        raise Exception("Identifier column (3) was moved! Logic error in col_map.")
    
    wb = openpyxl.load_workbook(file_path, data_only=False) 
    if "Last 6_DNT" not in wb.sheetnames:
        raise ValueError("Sheet 'Last 6_DNT' not found!")
    ws_last6 = wb["Last 6_DNT"]

    if "Normalization_DNT" in wb.sheetnames:
        wb.remove(wb["Normalization_DNT"])
    pre_group_index = wb.sheetnames.index("Group_DNT")
    ws_group = wb.create_sheet("Normalization_DNT", pre_group_index)

    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
            
    try:
        ws_last6.sheet_view.tabSelected = True
        ws_group.sheet_view.tabSelected = False
    except Exception:
        pass
        
    # --- DISTINCT FILL COLORS ---
    carb_fill = _make_fill("DAE9F8")   # Classic Blue
    co2_fill = _make_fill("E2EFDA")    # Soft Mint Green
    yield_fill = _make_fill("FFE0BF")  # Orange White
    
    dark_fill = _make_fill("808080")
    gray_fill = _make_fill("E7E7E7")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    color_fonts = {
        "nbs18": Font(color="FF0000"),
        "nbs19": Font(color="000080"),
        "iaea603": Font(color="008000"),
        "lsvec": Font(color="3399FF"),
    }
    
    MAX_GROUP_COL = dest_col - 1
    
    # DUAL TOP BLOCKS: Original Carbonate (Box Offset 0) & Conditional CO2 
    carb_bottom, _, carb_slope_info, carb_mat_map = draw_blue_box_structure(ws_group, 0, "Carbonate", carb_fill)
    
    if calc_co2:
        co2_bottom, _, co2_slope_info, co2_mat_map = draw_blue_box_structure(ws_group, CO2_TOP_SHIFT, "CO2", co2_fill)
        blue_box_bottom = max(carb_bottom, co2_bottom)
    else:
        co2_slope_info = []
        co2_mat_map = {}
        blue_box_bottom = carb_bottom
    
    # ---- INJECT YIELD CALCULATIONS TABLE INTO GAP DYNAMICALLY ----
    if calc_yield:
        draw_yield_table(ws_group, start_col=YIELD_START_COL, box_fill=yield_fill, num_yield_groups=num_yield_groups)
    
    header_row = blue_box_bottom + 4 + HEADER_ROW_SHIFT + EXTRA_VERT_SHIFT
    
    draw_upper_boxes(ws_group, header_row, carb_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"), 0, "Carbonate")
    if calc_co2:
        draw_upper_boxes(ws_group, header_row, co2_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"), CO2_OUT_SHIFT, "CO2")
    
    for row in ws_group.iter_rows(min_row=1, max_row=blue_box_bottom, min_col=1, max_col=MAX_GROUP_COL):
        for cell in row:
            cell.fill = carb_fill
            
    if calc_co2:
        for row in ws_group.iter_rows(min_row=1, max_row=blue_box_bottom, min_col=1 + CO2_TOP_SHIFT, max_col=16 + CO2_TOP_SHIFT):
            for cell in row:
                cell.fill = co2_fill
            
    headers = []
    first_row_cells = list(ws_last6[1]) if ws_last6.max_row >= 1 else []
    
    for src_col, dest_col_idx in col_map.items():
        src_cell = first_row_cells[src_col - 1] if src_col <= len(first_row_cells) else None
        header_val = src_cell.value if src_cell else None
        headers.append(header_val)
        dest_cell = ws_group.cell(row=header_row, column=dest_col_idx, value=header_val)
        if src_cell:
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.font = copy(src_cell.font)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)

    # --- INLINE HELPER FOR YIELD HEADERS ---
    red_if = InlineFont(color='00FF0000', b=True)
    blue_if = InlineFont(color='000000FF', b=True)
    green_if = InlineFont(color='008000', b=True)
    black_if = InlineFont(color='000000', b=True)
    
    def get_local_yield_rich_text(group_list):
        parts = []
        for item in group_list:
            clean = str(item).strip().upper()
            font = black_if
            if "18" in clean: font = red_if
            elif "19" in clean: font = blue_if
            elif "603" in clean: font = green_if
            parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
        return create_rich_text(parts)

    # --- THEORETICAL & YIELD HEADERS CREATION (MULTI-TIERED) ---
    theo_col = YIELD_START_COL + 1
    
    if calc_yield:
        if num_yield_groups > 0:
            yield_calc_end = theo_col + (num_yield_groups * 4) 
            thick = Side(style="thick")
            
            # 1. Draw the main box including Theoretical (increased height by 1)
            for r in range(header_row - 3, header_row + 1):
                for c in range(theo_col, yield_calc_end + 1):
                    cell = ws_group.cell(row=r, column=c)
                    cell.fill = yield_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Outer thick borders for the whole box
                    top = thick if r == header_row - 3 else None
                    bottom = thick if r == header_row else None
                    left = thick if c == theo_col else None
                    right = thick if c == yield_calc_end else None
                    cell.border = Border(top=top, bottom=bottom, left=left, right=right)

            # 2. "Normalized Yield" Title (merged horizontally)
            ws_group.merge_cells(start_row=header_row - 3, start_column=theo_col, end_row=header_row - 3, end_column=yield_calc_end)
            ny_cell = ws_group.cell(row=header_row - 3, column=theo_col, value="Normalized Yield")
            ny_cell.font = Font(bold=True)
            
            # 3. "Theoretical" Header (no vertical merge)
            theo_header = ws_group.cell(row=header_row, column=theo_col, value="Theoretical")
            theo_header.font = Font(bold=True, color="000000")
            
            # 4. Populate each Yield Group
            for i in range(num_yield_groups):
                calc_col = theo_col + 2 + (i * 4)
                yield_pct_col = calc_col + 2
                
                # "Calculated" and "Yield (%)" headers
                c_head = ws_group.cell(row=header_row - 1, column=calc_col, value="Calculated")
                c_head.font = Font(bold=True)
                
                y_head = ws_group.cell(row=header_row - 1, column=yield_pct_col, value="Yield (%)")
                y_head.font = Font(bold=True)
                
                # Reference Materials underneath (merged across the 3 columns)
                rt = get_local_yield_rich_text(yield_slope_groups[i])
                ws_group.merge_cells(start_row=header_row, start_column=calc_col, end_row=header_row, end_column=yield_pct_col)
                rt_cell = ws_group.cell(row=header_row, column=calc_col)
                rt_cell.value = rt
                
                # Ensure column widths are spacious
                ws_group.column_dimensions[get_column_letter(calc_col)].width = 15
                ws_group.column_dimensions[get_column_letter(yield_pct_col)].width = 15
        else:
            # Fallback if no yield groups are configured
            theo_header = ws_group.cell(row=header_row, column=theo_col, value="Theoretical")
            theo_header.fill = yield_fill
            theo_header.font = Font(bold=True, color="000000")
            theo_header.alignment = Alignment(horizontal="center", vertical="center")

    data_rows_with_index = []
    for row_idx, row_cells in enumerate(ws_last6.iter_rows(min_row=2, max_col=MAX_SOURCE_COL), start=2):
        row_values_filtered = []
        source_cell_formats = {}
        
        for src_col, cell in enumerate(row_cells, start=1):
            if src_col in col_map:
                row_values_filtered.append(cell.value)
                source_cell_formats[src_col] = {
                    'format': copy(cell.number_format),
                    'font': copy(cell.font), 
                    'fill': copy(cell.fill), 
                    'alignment': copy(cell.alignment),
                    'border': copy(cell.border)
                }
        
        if any(row_values_filtered):
            row_values_filtered = row_values_filtered + [None] * (MAX_GROUP_COL - len(row_values_filtered))
            data_rows_with_index.append((row_idx, tuple(row_values_filtered), source_cell_formats))

    groups = {}

    for original_row_idx, r_values, r_formats in data_rows_with_index:
        ident = r_values[col_identifier1 - 1]
        base = extract_sample_base(ident)
        norm = _normalize_text(base)
        
        if norm not in groups:
            groups[norm] = {"base": base, "rows": []}
        groups[norm]["rows"].append((original_row_idx, r_values, r_formats)) 
        
    for g in groups.values():
        g["rows"].sort(key=lambda item: extract_run_number(item[1][col_identifier1 - 1]))
        
    carb_materials = settings.get_setting("REFERENCE_MATERIALS", sub_key="Carbonate") or []

    # --- Fetch UI settings for manual Drag & Drop mapping ---
    active_refs = settings.get_setting("ACTIVE_REFERENCES")
    active_samples = settings.get_setting("ACTIVE_SAMPLES")

    def is_reference_group(base_name):
        if not base_name: return False
        raw_text = str(base_name).strip()
        text_clean = re.sub(r'[\s\-_]+', '', raw_text.upper())
        text_no_std = text_clean.replace("STD", "")
        
        # 1. Trust the manual UI Drag & Drop lists completely
        if active_refs is not None and raw_text in active_refs:
            return True
        if active_samples is not None and raw_text in active_samples:
            return False
            
        # 2. Hardcoded exceptions (CO2/HeCO2)
        if text_clean.startswith("CO2") or text_clean.startswith("HECO2") or text_clean.startswith("C02") or text_clean.startswith("HEC02"):
            return True
            
        # 3. Check the global Advanced Settings table (carb_materials)
        for mat in carb_materials:
            std_name = mat.get("col_c")
            if not std_name: continue
            
            std_clean = re.sub(r'[\s\-_]+', '', str(std_name).upper())
            std_no_std = std_clean.replace("STD", "")
            
            if std_clean and std_clean in text_clean:
                return True
            if std_no_std and len(std_no_std) >= 4 and std_no_std in text_no_std:
                return True
                
        return False

    ref_groups = []
    other_groups = []
    for norm, g in groups.items():
        if is_reference_group(g["base"]):
            ref_groups.append((norm, g))
        else:
            other_groups.append((norm, g))

    present_refs_norm = {norm for norm, _ in ref_groups}
    has_iaea_603 = 'iaea603' in present_refs_norm
            
    current_row = header_row + 1
    n_arag_re = re.compile(r"\bn\.?\s*arag\b", flags=re.IGNORECASE)
    
    yield_mat_row_map = {}

    def _write_sample_output_cells(ws, r, is_arag, has_iaea_603, new_R_col, new_U_col, output_format, slope_info, box_offset=0, output_offset=0):
        num_groups = len(slope_info)
        if num_groups == 0:
            return 27 + output_offset
        
        c_start = 19 + output_offset
        o_calc_start = c_start + num_groups + 1
        o_arag_start = o_calc_start + num_groups + 1
        
        box1_start = c_start
        box1_end = o_arag_start + num_groups - 1
        
        box2_start = box1_end + 2
        vsmow_calc_start = box2_start
        vsmow_arag_start = vsmow_calc_start + num_groups + 1
        
        col_AA_approx = vsmow_arag_start + num_groups - 1
        
        green_bold = Font(color="008000", bold=True)
        
        c_slope_col = get_column_letter(11 + box_offset)
        o_slope_col = get_column_letter(14 + box_offset)
        
        for i in range(num_groups):
            s_row = slope_info[i]["slope_row"]
            i_row = slope_info[i]["intercept_row"]
            
            c_formula = f'=IFERROR(${c_slope_col}${s_row}*{get_column_letter(new_R_col)}{r}+${c_slope_col}${i_row},"")'
            o_formula = f'=IFERROR(${o_slope_col}${s_row}*{get_column_letter(new_U_col)}{r}+${o_slope_col}${i_row},"")'
            
            c_col = c_start + i
            o_calc_col = o_calc_start + i
            o_arag_col = o_arag_start + i
            vsmow_calc_col = vsmow_calc_start + i
            vsmow_arag_col = vsmow_arag_start + i
            
            cell_c = ws.cell(row=r, column=c_col, value=c_formula)
            cell_c.number_format = output_format
            
            if is_arag:
                ws.cell(row=r, column=o_calc_col, value=None)
                ws.cell(row=r, column=vsmow_calc_col, value=None)
                
                cell_o_arag = ws.cell(row=r, column=o_arag_col, value=o_formula)
                cell_o_arag.font = Font(bold=True)
                cell_o_arag.number_format = output_format
                
                vsmow_formula = f'=IFERROR((1.03092*{get_column_letter(o_arag_col)}{r})+30.92,"")'
                cell_vsmow_arag = ws.cell(row=r, column=vsmow_arag_col, value=vsmow_formula)
                cell_vsmow_arag.font = Font(bold=True)
                cell_vsmow_arag.number_format = output_format
                
                for col_idx in (c_col, o_arag_col, vsmow_arag_col):
                    ws.cell(row=r, column=col_idx).font = green_bold
            else:
                ws.cell(row=r, column=o_arag_col, value=None)
                ws.cell(row=r, column=vsmow_arag_col, value=None)
                
                cell_o_calc = ws.cell(row=r, column=o_calc_col, value=o_formula)
                cell_o_calc.number_format = output_format
                
                vsmow_formula = f'=IFERROR((1.03092*{get_column_letter(o_calc_col)}{r})+30.92,"")'
                cell_vsmow_calc = ws.cell(row=r, column=vsmow_calc_col, value=vsmow_formula)
                cell_vsmow_calc.font = Font(bold=True)
                cell_vsmow_calc.number_format = output_format
                
                for col_idx in (c_col, o_calc_col, vsmow_calc_col):
                    original_font = copy(ws.cell(row=r, column=col_idx).font)
                    original_font.bold = True
                    ws.cell(row=r, column=col_idx).font = original_font
                    
        return col_AA_approx

    def write_group(norm, g, is_reference=True, has_iaea_603=False):
        nonlocal current_row
        base_name = _normalize_text(g["base"])
        rows_data = g["rows"]
        font_color = color_fonts.get(base_name)
        
        row_values_list = [item[1] for item in rows_data]
        valid_indices = []
        if base_name in ("co2", "heco2"):
            valid_indices = _get_valid_co2_rows(row_values_list, col_identifier1)
            
        row_map = [] 
        
        c_vals = []; o_vals = [] 
        valid_run_row_indices = [] 

        for i, (source_row_idx, row_values_filtered, source_cell_formats) in enumerate(rows_data):
            excel_row = current_row
            
            for dest_col_idx in range(1, MAX_GROUP_COL + 1):
                src_col = next(s for s, d in col_map.items() if d == dest_col_idx)
                val = row_values_filtered[dest_col_idx - 1]
                dest_cell = ws_group.cell(row=excel_row, column=dest_col_idx, value=val)
                
                original_format = source_cell_formats.get(src_col, {})
                if original_format.get('format'): dest_cell.number_format = original_format['format']
                if original_format.get('font'): dest_cell.font = original_format['font']
                if original_format.get('fill'): dest_cell.fill = original_format['fill']
                if original_format.get('alignment'): dest_cell.alignment = original_format['alignment']
                if original_format.get('border'): dest_cell.border = original_format['border']

                if font_color and dest_col_idx == col_identifier1:
                    new_font = copy(dest_cell.font)
                    new_font.color = font_color.color
                    dest_cell.font = new_font
                    
            if base_name in ("co2", "heco2") and i in valid_indices:
                for col in range(1, MAX_GROUP_COL + 1):
                    ws_group.cell(row=excel_row, column=col).fill = gray_fill
                    
            ident_val = str(ws_group.cell(row=excel_row, column=col_identifier1).value or "")
            is_arag = n_arag_re.search(ident_val)
            
            _write_sample_output_cells(ws_group, excel_row, bool(is_arag), has_iaea_603, 11, 14, '0.00', carb_slope_info, 0, 0)
            
            if calc_co2:
                _write_sample_output_cells(ws_group, excel_row, bool(is_arag), has_iaea_603, 11, 14, '0.00', co2_slope_info, CO2_TOP_SHIFT, CO2_OUT_SHIFT)
            
            # --- THEORETICAL & YIELD FORMULA INJECTION ---
            if calc_yield and base_name not in ("co2", "heco2"):
                col_an = get_column_letter(YIELD_START_COL + 7)
                
                target_row_val = ref_compound_row if is_reference else samp_compound_row
                ref_cell = f"${col_an}${target_row_val}"
                
                theo_cell = ws_group.cell(row=excel_row, column=theo_col, value=f'=IF(ISNUMBER(G{excel_row}), G{excel_row}*{ref_cell}, "")')
                theo_cell.number_format = '0.00E+00'
                
                for grp_idx in range(num_yield_groups):
                    calc_col = theo_col + 2 + (grp_idx * 4)
                    yield_pct_col = calc_col + 2
                    
                    slope_cell = f"${get_column_letter(YIELD_START_COL + 5)}${20 + grp_idx*3}"
                    int_cell = f"${get_column_letter(YIELD_START_COL + 5)}${21 + grp_idx*3}"
                    
                    calc_cell = ws_group.cell(row=excel_row, column=calc_col, value=f'=IFERROR(({slope_cell}*Q{excel_row})+{int_cell},"")')
                    calc_cell.number_format = '0.00E+00'
                    
                    yp_cell = ws_group.cell(row=excel_row, column=yield_pct_col, value=f'=IFERROR(({get_column_letter(calc_col)}{excel_row}/{get_column_letter(theo_col)}{excel_row})*100,"")')
                    yp_cell.number_format = '0.00'

            should_process = True
            if base_name in ("co2", "heco2") and i not in valid_indices:
                should_process = False
                c_vals.append(None); o_vals.append(None)
            
            if should_process:
                valid_run_row_indices.append(i)
                cv = row_values_filtered[11-1] 
                ov = row_values_filtered[14-1] 
                try: cv = float(cv)
                except: cv = None
                try: ov = float(ov)
                except: ov = None
                c_vals.append(cv); o_vals.append(ov)
                
                if base_name not in yield_mat_row_map:
                    yield_mat_row_map[base_name] = []
                yield_mat_row_map[base_name].append(excel_row)

            row_map.append(excel_row)
            current_row += 1
            
        if is_reference or len(rows_data) > 1:
            valid_c_nums = [v for v in c_vals if v is not None]
            valid_o_nums = [v for v in o_vals if v is not None]

            mean_c = statistics.mean(valid_c_nums) if len(valid_c_nums) > 1 else 0
            stdev_c = statistics.stdev(valid_c_nums) if len(valid_c_nums) > 1 else 0
            mean_o = statistics.mean(valid_o_nums) if len(valid_o_nums) > 1 else 0
            stdev_o = statistics.stdev(valid_o_nums) if len(valid_o_nums) > 1 else 0

            c_up, c_low = mean_c + (outlier_sigma * stdev_c), mean_c - (outlier_sigma * stdev_c)
            o_up, o_low = mean_o + (outlier_sigma * stdev_o), mean_o - (outlier_sigma * stdev_o)

            all_runs_c = []; all_runs_o = []
            final_runs_rows = []

            c_col_let = "K"; o_col_let = "N"

            for i in valid_run_row_indices:
                r_num = row_map[i]
                vc = c_vals[i]; vo = o_vals[i]

                if vc is not None: all_runs_c.append(f"{c_col_let}{r_num}")
                if vo is not None: all_runs_o.append(f"{o_col_let}{r_num}")

                is_c_out = (vc > c_up or vc < c_low) if (vc is not None and len(valid_c_nums) > 2) else False
                is_o_out = (vo > o_up or vo < o_low) if (vo is not None and len(valid_o_nums) > 2) else False

                exclude_c = False; exclude_o = False
                if exclusion_mode == "Exclude Row":
                    if is_c_out or is_o_out: exclude_c = True; exclude_o = True
                else:
                    if is_c_out: exclude_c = True
                    if is_o_out: exclude_o = True
                
                if not exclude_c or not exclude_o: final_runs_rows.append(r_num)
                
                if exclude_c and vc is not None: ws_group.cell(row=r_num, column=11).font = strike_font
                if exclude_o and vo is not None: ws_group.cell(row=r_num, column=14).font = strike_font
        
            summary_font = font_color or Font(bold=True)
            if is_reference and font_color:
                summary_font = copy(font_color); summary_font.bold = True
            
            data_fmt = get_summary_num_format(base_name); count_fmt = '0'

            row_all = current_row
            ws_group.cell(row=row_all, column=10, value="--").font = Font(bold=True) 
            for col, txt in zip([11, 12, 13, 14, 15, 16], ["Average", "Stdev", "Count", "Average", "Stdev", "Count"]):
                c = ws_group.cell(row=row_all, column=col, value=txt)
                c.font = summary_font; c.alignment = Alignment(horizontal="right")
                
            if calc_yield and base_name not in ("co2", "heco2"):
                for grp_idx in range(num_yield_groups):
                    yield_pct_col = theo_col + 4 + (grp_idx * 4)
                    c = ws_group.cell(row=row_all, column=yield_pct_col, value="Average")
                    c.font = summary_font; c.alignment = Alignment(horizontal="right")

            row_all_calc = row_all + 1
            if all_runs_c:
                rng_c = ",".join(all_runs_c)
                ws_group.cell(row=row_all_calc, column=11, value=f"=AVERAGE({rng_c})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=12, value=f"=STDEV({rng_c})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=13, value=f"=COUNT({rng_c})").number_format = count_fmt
            if all_runs_o:
                rng_o = ",".join(all_runs_o)
                ws_group.cell(row=row_all_calc, column=14, value=f"=AVERAGE({rng_o})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=15, value=f"=STDEV({rng_o})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=16, value=f"=COUNT({rng_o})").number_format = count_fmt
            
            for c in range(11, 17): ws_group.cell(row=row_all_calc, column=c).font = summary_font
            
            if calc_yield and base_name not in ("co2", "heco2"):
                all_runs_rows = [row_map[i] for i in valid_run_row_indices]
                for grp_idx in range(num_yield_groups):
                    yield_pct_col = theo_col + 4 + (grp_idx * 4)
                    y_col_let = get_column_letter(yield_pct_col)
                    y_range_str = get_excel_range(y_col_let, all_runs_rows)
                    if y_range_str:
                        ws_group.cell(row=row_all_calc, column=yield_pct_col, value=f'=IFERROR(AVERAGE({y_range_str}),"")').number_format = '0.00'
                        ws_group.cell(row=row_all_calc, column=yield_pct_col).font = summary_font

            is_arag_group = n_arag_re.search(g["base"])
            _write_sample_output_cells(ws_group, row_all_calc, bool(is_arag_group), has_iaea_603, 11, 14, '0.00', carb_slope_info, 0, 0)
            if calc_co2:
                _write_sample_output_cells(ws_group, row_all_calc, bool(is_arag_group), has_iaea_603, 11, 14, '0.00', co2_slope_info, CO2_TOP_SHIFT, CO2_OUT_SHIFT)

            for col_rng in [range(19+0, 32+0), range(19+CO2_OUT_SHIFT, 32+CO2_OUT_SHIFT)]: 
                for c in col_rng:
                    if ws_group.cell(row=row_all_calc, column=c).value: 
                        ws_group.cell(row=row_all_calc, column=c).font = Font(bold=True)

            row_filt = row_all_calc + 1
            ws_group.cell(row=row_filt, column=10, value="Outlier Excl.").font = Font(bold=True) 
            for col, txt in zip([11, 12, 13, 14, 15, 16], ["Average", "Stdev", "Count", "Average", "Stdev", "Count"]):
                c = ws_group.cell(row=row_filt, column=col, value=txt)
                c.font = summary_font; c.alignment = Alignment(horizontal="right")

            row_filt_calc = row_filt + 1
            
            # Reconstruct filtered ranges for C and O specifically, or use final_runs_rows for Yield
            rng_c_filt = [f"{c_col_let}{r}" for r in final_runs_rows if f"{c_col_let}{r}" in all_runs_c]
            rng_o_filt = [f"{o_col_let}{r}" for r in final_runs_rows if f"{o_col_let}{r}" in all_runs_o]

            if rng_c_filt:
                rng_c_str = ",".join(rng_c_filt)
                ws_group.cell(row=row_filt_calc, column=11, value=f"=AVERAGE({rng_c_str})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=12, value=f"=STDEV({rng_c_str})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=13, value=f"=COUNT({rng_c_str})").number_format = count_fmt
            if rng_o_filt:
                rng_o_str = ",".join(rng_o_filt)
                ws_group.cell(row=row_filt_calc, column=14, value=f"=AVERAGE({rng_o_str})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=15, value=f"=STDEV({rng_o_str})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=16, value=f"=COUNT({rng_o_str})").number_format = count_fmt

            for c in range(11, 17): ws_group.cell(row=row_filt_calc, column=c).font = summary_font

            _write_sample_output_cells(ws_group, row_filt_calc, bool(is_arag_group), has_iaea_603, 11, 14, '0.00', carb_slope_info, 0, 0)
            if calc_co2:
                _write_sample_output_cells(ws_group, row_filt_calc, bool(is_arag_group), has_iaea_603, 11, 14, '0.00', co2_slope_info, CO2_TOP_SHIFT, CO2_OUT_SHIFT)

            for col_rng in [range(19+0, 32+0), range(19+CO2_OUT_SHIFT, 32+CO2_OUT_SHIFT)]: 
                for c in col_rng:
                    if ws_group.cell(row=row_filt_calc, column=c).value: 
                        ws_group.cell(row=row_filt_calc, column=c).font = Font(bold=True)

            current_row = row_filt_calc + 2 
        else:
            current_row += 3
            
    for norm, g in ref_groups:
        write_group(norm, g, is_reference=True)
        
    if ref_groups:
        current_row += 8
        divider_top_row = current_row 
        for _ in range(2):
            for col in range(1, 702):
                ws_group.cell(row=current_row, column=col).fill = dark_fill
            current_row += 1
            
        current_header_col = 1
        for src_col, dest_col_idx in col_map.items():
            src_cell = first_row_cells[src_col - 1] if src_col <= len(first_row_cells) else None
            
            dest_cell = ws_group.cell(row=current_row, column=current_header_col, value=headers[current_header_col - 1])
            
            if src_cell:
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.font = copy(src_cell.font)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)

        # BOTTOM YIELD BOX START
        if calc_yield:
            if num_yield_groups > 0:
                yield_calc_end = theo_col + (num_yield_groups * 4)
                thick = Side(style="thick")
                
                # Align perfectly with the Carbonate and CO2 lower boxes (which span from -3 to +1)
                bot_y_top = divider_top_row - 3
                bot_y_bottom = divider_top_row + 1
                
                # 1. Draw the main box
                for r in range(bot_y_top, bot_y_bottom + 1):
                    for c in range(theo_col, yield_calc_end + 1):
                        cell = ws_group.cell(row=r, column=c)
                        cell.fill = yield_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        top_border = thick if r == bot_y_top else None
                        bottom_border = thick if r == bot_y_bottom else None
                        left_border = thick if c == theo_col else None
                        right_border = thick if c == yield_calc_end else None
                        cell.border = Border(top=top_border, bottom=bottom_border, left=left_border, right=right_border)
                        
                # 2. "Normalized Yield" Title (merged horizontally)
                ws_group.merge_cells(start_row=bot_y_top, start_column=theo_col, end_row=bot_y_top, end_column=yield_calc_end)
                ny_cell = ws_group.cell(row=bot_y_top, column=theo_col, value="Normalized Yield")
                ny_cell.font = Font(bold=True)
                
                # 3. "Theoretical" Header (placed near the top, no vertical merge)
                theo_header_bottom = ws_group.cell(row=bot_y_top + 4, column=theo_col, value="Theoretical")
                theo_header_bottom.font = Font(bold=True, color="000000")
                
                # 4. Populate each Yield Group
                for i in range(num_yield_groups):
                    calc_col = theo_col + 2 + (i * 4)
                    yield_pct_col = calc_col + 2
                    
                    # "Calculated" and "Yield (%)" headers
                    c_head = ws_group.cell(row=bot_y_bottom - 1, column=calc_col, value="Calculated")
                    c_head.font = Font(bold=True)
                    
                    y_head = ws_group.cell(row=bot_y_bottom - 1, column=yield_pct_col, value="Yield (%)")
                    y_head.font = Font(bold=True)
                    
                    # Reference Materials underneath (merged across the 3 columns)
                    rt = get_local_yield_rich_text(yield_slope_groups[i])
                    ws_group.merge_cells(start_row=bot_y_bottom, start_column=calc_col, end_row=bot_y_bottom, end_column=yield_pct_col)
                    rt_cell = ws_group.cell(row=bot_y_bottom, column=calc_col)
                    rt_cell.value = rt
            else:
                theo_header_bottom = ws_group.cell(row=current_row, column=theo_col, value="Theoretical")
                theo_header_bottom.fill = yield_fill
                theo_header_bottom.font = Font(bold=True, color="000000")
                theo_header_bottom.alignment = Alignment(horizontal="center", vertical="center")
        # BOTTOM YIELD BOX END

        current_header_col += 1

        current_row += 1
        
        draw_lower_boxes(ws_group, divider_top_row, carb_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"), 0, "Carbonate")
        if calc_co2:
            draw_lower_boxes(ws_group, divider_top_row, co2_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"), CO2_OUT_SHIFT, "CO2")
        
    for norm, g in other_groups:
        write_group(norm, g, is_reference=False, has_iaea_603=has_iaea_603)
        
    populate_blue_box_math(ws_group, carb_slope_info, carb_mat_map, 0, "Carbonate")
    if calc_co2:
        populate_blue_box_math(ws_group, co2_slope_info, co2_mat_map, CO2_TOP_SHIFT, "CO2")
    
    # --- YIELD SLOPE & INTERCEPT DYNAMIC POPULATION ---
    if calc_yield:
        def get_group_rich_text(group_list):
            parts = [(black_if, "Used ")]
            for item in group_list:
                clean = str(item).strip().upper()
                font = black_if
                if "18" in clean: font = red_if
                elif "19" in clean: font = blue_if
                elif "603" in clean: font = green_if
                parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
            return create_rich_text(parts)
            
        yield_slope_row = 20 
        col_AI = YIELD_START_COL + 2
        col_AJ = YIELD_START_COL + 3
        col_AL = YIELD_START_COL + 5
        
        for group in yield_slope_groups:
            group_rows = []
            for mat in group:
                mat_clean = re.sub(r'[\s\-_]+', '', mat.upper()).replace("STD", "")
                for key, rows in yield_mat_row_map.items():
                    if mat_clean in key.upper() or key.upper() in mat_clean:
                        group_rows.extend(rows)
                        
            if len(group_rows) >= 2:
                ws_group.cell(row=yield_slope_row, column=col_AI).value = get_group_rich_text(group)
                
                ws_group.cell(row=yield_slope_row, column=col_AJ, value="slope").font = Font(bold=True)
                ws_group.cell(row=yield_slope_row+1, column=col_AJ, value="intercept").font = Font(bold=True)
                
                range_y = get_excel_range(get_column_letter(theo_col), group_rows)
                range_x = get_excel_range('Q', group_rows)
                
                cell_slope = ws_group.cell(row=yield_slope_row, column=col_AL, value=f'=SLOPE({range_y},{range_x})')
                cell_slope.number_format = '0.00E+00'
                cell_slope.font = Font(bold=True)
                
                cell_int = ws_group.cell(row=yield_slope_row+1, column=col_AL, value=f'=INTERCEPT({range_y},{range_x})')
                cell_int.number_format = '0.00E+00'
                cell_int.font = Font(bold=True)
                
                yield_slope_row += 3 
            
    start_gray_row = header_row + 1 
    max_sheet_row = ws_group.max_row + 50
    
    for row in range(start_gray_row, max_sheet_row + 1):
        for col in (11, 14): 
            ws_group.cell(row=row, column=col).fill = gray_fill
            
    ws_group.column_dimensions["A"].width = 13 
    ws_group.column_dimensions["C"].width = 22 
    ws_group.column_dimensions["H"].width = 15 
    ws_group.column_dimensions["J"].width = 16 

    offsets = [0]
    if calc_co2: offsets.append(CO2_TOP_SHIFT)
    
    for offset in offsets:
        ws_group.column_dimensions[get_column_letter(3 + offset)].width = 22
        ws_group.column_dimensions[get_column_letter(6 + offset)].width = 12
        ws_group.column_dimensions[get_column_letter(7 + offset)].width = 12
        ws_group.column_dimensions[get_column_letter(10 + offset)].width = 16
        ws_group.column_dimensions[get_column_letter(11 + offset)].width = 12
        ws_group.column_dimensions[get_column_letter(14 + offset)].width = 12
        
    out_offsets = [0]
    if calc_co2: out_offsets.append(CO2_OUT_SHIFT)
    
    for output_offset in out_offsets:
        for output_col in range(19 + output_offset, 30 + output_offset):
            ws_group.column_dimensions[get_column_letter(output_col)].width = 12
            
    if calc_yield:
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 1)].width = 15 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 2)].width = 15 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 3)].width = 15 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 5)].width = 15 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 7)].width = 15 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 9)].width = 10 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 10)].width = 10 
        ws_group.column_dimensions[get_column_letter(YIELD_START_COL + 11)].width = 10 

    max_data_row = ws_group.max_row
    threshold_str = str(stdev_threshold)
    
    ws_group.conditional_formatting.add(
        f"L{start_gray_row}:L{max_data_row}",
        FormulaRule(formula=[f'=AND(ISNUMBER(L{start_gray_row}), L{start_gray_row} > {threshold_str})'], fill=red_fill)
    )

    ws_group.conditional_formatting.add(
        f"O{start_gray_row}:O{max_data_row}",
        FormulaRule(formula=[f'=AND(ISNUMBER(O{start_gray_row}), O{start_gray_row} > {threshold_str})'], fill=red_fill)
    )

    # Dynamic Freeze Pane (Conditional)
    if settings.get_setting("ENABLE_FREEZE_PANE") is not False:
        ws_group.freeze_panes = f'B{start_gray_row}'

    embed_settings_popup(ws_group, "A2")

    wb.save(file_path)
    print(f"✅ Step 6: Normalization completed on {file_path}")